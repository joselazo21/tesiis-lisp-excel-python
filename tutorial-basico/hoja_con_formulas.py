"""
hoja_con_formulas.py — Framework de Generación de Excel
========================================================

Genera archivos .xlsx completos con datos, fórmulas, estilos, bordes,
formato condicional y celdas fusionadas a partir de una configuración
declarativa en Python (diccionarios).

Puede ser alimentado por cualquier generador externo: Common Lisp, JSON,
YAML, bases de datos, etc.

Dependencias
------------
    pip install openpyxl

Uso mínimo
----------
    from hoja_con_formulas import generate_excel

    generate_excel({
        "sheets": [{
            "title": "Mi Hoja",
            "data": [["A", "B"], [1, 2]],
            "table_borders": True
        }]
    }, "salida.xlsx")

Alias de compatibilidad
-----------------------
    generar_excel_personalizado = generate_excel


═══════════════════════════════════════════════════════════════════════
ARQUITECTURA: Pipeline de procesamiento
═══════════════════════════════════════════════════════════════════════

generate_excel() recorre cada hoja y aplica estas etapas en orden:

    1. _process_headers()          → Encabezados (opcional, usa ws.append)
    2. _process_data()             → Filas de datos (ws.append)
    3. _process_formulas()         → Fórmulas simples por coordenada
    4. _inject_fernando_formulas() → Fórmulas complejas (cross-sheet)
    5. _process_column_widths()    → Ancho de columnas
    6. _process_legacy_conditional_formatting() → Formato condicional legacy
    7. _process_cell_scaling()     → Escalar filas/columnas por factor
    8. _process_table_borders()    → Bordes (uniformes o por bloques)
    9. _process_range_styles()     → Colores de fondo / fuente por rango
   10. _process_merge_ranges()     → Celdas fusionadas
   11. _process_conditional_format_rules() → Formato condicional avanzado

IMPORTANTE: Los estilos de rango (paso 9) se aplican DESPUÉS de los datos
(paso 2), por lo que sobrescriben cualquier formato previo. El orden es
deliberado: datos → fórmulas → bordes → colores → fusiones → condicional.


═══════════════════════════════════════════════════════════════════════
REFERENCIA DE CONFIGURACIÓN DE HOJA
═══════════════════════════════════════════════════════════════════════

Cada elemento de config["sheets"] acepta las siguientes claves:

───────────────────────────────────────────────────────────────────────
OBLIGATORIAS
───────────────────────────────────────────────────────────────────────

title : str
    Nombre de la hoja en el libro Excel.

───────────────────────────────────────────────────────────────────────
DATOS
───────────────────────────────────────────────────────────────────────

data : list[list]
    Filas de datos. Cada sublista es una fila.
    Se agregan con ws.append(), por lo que la primera sublista va en la
    fila 1 (o fila 2 si se usó "headers").

    Ejemplo:
        "data": [
            ["", "", "1", "2", "3"],
            ["", "Lunes", "Aula 1", "Aula 2", "Aula 3"],
            ["", "1ro", "", "", ""],
        ]

headers : list[str]  (opcional)
    Fila de encabezados. Se inserta ANTES de data con ws.append().
    Se estiliza con header_style.

    Ejemplo:
        "headers": ["ID", "Nombre", "Valor"]

───────────────────────────────────────────────────────────────────────
FÓRMULAS SIMPLES
───────────────────────────────────────────────────────────────────────

formulas : list[dict]  (opcional)
    Fórmulas o valores en celdas específicas, referenciadas por
    coordenadas numéricas (row, col) 1-based.

    Cada dict tiene:
        row   : int   — número de fila (1-based)
        col   : int   — número de columna (1-based)
        value : str   — fórmula (empieza con "=") o valor literal

    Ejemplo:
        "formulas": [
            {"row": 5, "col": 13, "value": "=COUNTIF($C$4:$G$15,I5)"},
            {"row": 5, "col": 12, "value": "=K5-M5"},
            {"row": 11, "col": 12, "value": "Total:"},
        ]

───────────────────────────────────────────────────────────────────────
FÓRMULAS COMPLEJAS (Cross-Sheet)
───────────────────────────────────────────────────────────────────────

fernando_formulas : list[dict]  (opcional)
    Fórmulas complejas referenciadas por notación Excel (A1, C4, etc.).
    Diseñado para fórmulas que cruzan múltiples hojas.

    Cada dict tiene:
        cell    : str — referencia Excel ("C4")
        formula : str — la fórmula (puede estar en español o inglés)
        # o alternativamente:
        excel   : str — la fórmula (se acepta "formula" o "excel")

    El motor normaliza automáticamente:
        • Referencias ODS [$C111.$C$9] → Excel C111!$C$9
        • Referencias con punto $C111.$D5 → C111!$D5
        • Funciones en español SI() → IF()
        • Funciones en español SUSTITUIR() → SUBSTITUTE()
        • Funciones en español ESPACIOS() → TRIM()
        • CONCATENAR(a;b;c) → (a & b & c)
        • Separadores ; → ,

    Ejemplo:
        "fernando_formulas": [
            {
                "cell": "C4",
                "formula": "=SUBSTITUTE(TRIM(CONCAT("
                    "IF(C111!$C$5=C$3,C111!$B$1 & \" \", \"\"),"
                    "IF(C121!$C$5=C$3,C121!$B$1 & \" \", \"\"))), "
                    "\" \", \",\")"
            }
        ]

───────────────────────────────────────────────────────────────────────
COLUMNAS
───────────────────────────────────────────────────────────────────────

column_widths : dict[int, float]  (opcional)
    Ancho de columnas por índice (1-based).

    Ejemplo:
        "column_widths": {1: 14, 2: 14, 3: 20}

        # O con comprensión:
        "column_widths": {i: 12 for i in range(1, 13)}

cell_size : float  (opcional)
    Factor de escala para TODAS las celdas de la hoja.
    2.0 = doble tamaño. Afecta alto de fila y ancho de columna.

    Ejemplo:
        "cell_size": 1.5

───────────────────────────────────────────────────────────────────────
ESTILOS DE ENCABEZADO
───────────────────────────────────────────────────────────────────────

header_style : dict  (opcional)
    Estilo aplicado a la fila de "headers".

    Claves soportadas:
        bold     : bool   — negrita
        color    : str    — color de fuente (RGB 6 chars)
        align    : str    — 'center', 'left', 'right'
        bg_color : str    — color de fondo (RGB 6 chars)

    Ejemplo:
        "header_style": {"bold": True, "align": "center", "bg_color": "E6B8AF"}

───────────────────────────────────────────────────────────────────────
ESTILOS POR RANGO
───────────────────────────────────────────────────────────────────────

range_styles : list[dict]  (opcional)
    Aplica estilos a rangos específicos de celdas.

    Cada dict tiene:
        range : str  — rango Excel ("B3:L3")
        style : dict — mismas claves que header_style

    Ejemplo:
        "range_styles": [
            {"range": "B3:L3",  "style": {"bg_color": "E6B8AF"}},
            {"range": "B11:L11", "style": {"bg_color": "E6B8AF"}},
            {"range": "I3:I10", "style": {"bg_color": "A9D18E"}},
        ]

───────────────────────────────────────────────────────────────────────
BORDES
───────────────────────────────────────────────────────────────────────

table_borders : bool  (opcional)
    Activa el sistema de bordes.

border_color : str  (opcional, default "000000")
    Color del borde en hex RGB.

border_style : str  (opcional, default "thin")
    Estilo del borde: "thin", "medium", "thick".

border_step : int  (opcional, default 1)
    Tamaño base de bloque para bordes agrupados.

table_ranges : list[str]  (opcional)
    Lista de rangos Excel donde aplicar bordes.
    Si no se especifica, se aplica a toda la hoja usada.

    Ejemplo:
        "table_ranges": ["B3:L9", "B11:L17", "B19:L25"]

table_block_sizes : list[dict]  (opcional)
    Configura bordes por bloques para rangos específicos.
    Útil cuando las celdas están agrupadas (ej: turnos de 2 filas).

    Cada dict tiene:
        range         : str   — rango Excel
        row_step      : int   — altura del bloque en filas
        col_step      : int   — ancho del bloque en columnas
        skip_first_row : bool — saltar primera fila del bloque
        skip_first_col : bool — saltar primera columna del bloque

    Ejemplo (turnos de 2 filas):
        "table_block_sizes": [
            {"range": "B4:B15", "row_step": 2, "col_step": 1},
            {"range": "C4:G15", "row_step": 2, "col_step": 1},
        ]

    Sin table_block_sizes, los bordes se aplican celda por celda.

───────────────────────────────────────────────────────────────────────
CELDA FUSIONADAS
───────────────────────────────────────────────────────────────────────

merge_ranges : list[str]  (opcional)
    Lista de rangos a fusionar. El contenido se centra automáticamente.

    Ejemplo:
        "merge_ranges": ["B4:B5", "B6:B7", "B8:B9"]

───────────────────────────────────────────────────────────────────────
FORMATO CONDICIONAL (AVANZADO)
───────────────────────────────────────────────────────────────────────

conditional_format_rules : list[dict]  (opcional)
    Reglas de formato condicional con plantillas de fórmula.

    Cada dict tiene:
        tipo    : str  — tipo de aplicación (ver abajo)
        rango   : str  — rango Excel donde aplicar
        formula : str  — plantilla de fórmula con placeholders
        color   : str  — color de relleno cuando se cumple
        stop_if_true : bool  — detener evaluación si cumple (default True)

        # Solo para tipo "pares_con_siguiente":
        aplicar_a : str  — "siguiente" o "actual"

    Placeholders en la fórmula:
        {celda}          — referencia actual (ej: "C4")
        {celda_siguiente} — referencia fila siguiente (ej: "C5")
        {fila}           — número de fila (ej: "4")
        {columna}        — letra de columna (ej: "C")

    Tipos de regla:
        "rango"          — aplica a cada celda del rango
        "filas_pares"    — aplica a filas 1, 3, 5... relativas al rango
        "filas_impares"  — aplica a filas 2, 4, 6... relativas al rango
        "pares_con_siguiente" — evalúa par+siguiente, aplica a una

    Ejemplo:
        "conditional_format_rules": [
            {
                "tipo": "filas_pares",
                "rango": "C4:G15",
                "formula": 'AND({celda}<>"", COUNTIF($I4:I10,{celda})=0)',
                "color": "F4A460"
            },
            {
                "tipo": "pares_con_siguiente",
                "rango": "C4:G15",
                "formula": 'AND({celda}<>"", {celda_siguiente}="")',
                "color": "FF0000",
                "aplicar_a": "siguiente"
            },
        ]

───────────────────────────────────────────────────────────────────────
FORMATO CONDICIONAL (LEGACY)
───────────────────────────────────────────────────────────────────────

conditional_formatting : list[dict]  (opcional)
    Formato condicional simple (formato legacy).

    Cada dict tiene:
        range     : str  — rango Excel
        rule      : str  — fórmula (sin placeholders)
        fill_color: str  — color de relleno

    Ejemplo:
        "conditional_formatting": [
            {
                "range": "A1:A10",
                "rule": 'A1>100',
                "fill_color": "FF0000"
            }
        ]


═══════════════════════════════════════════════════════════════════════
API PÚBLICA
═══════════════════════════════════════════════════════════════════════

generate_excel(config: dict, filename: str) -> None
    Función principal. Genera el archivo Excel.

generar_excel_personalizado(config: dict, filename: str) -> None
    Alias de generate_excel para compatibilidad.

───────────────────────────────────────────────────────────────────────
Funciones reutilizables (uso directo opcional)
───────────────────────────────────────────────────────────────────────

parse_range(range_str) -> tuple
    Parsea "C4:G17" → (3, 4, 7, 17)  # (min_col, min_row, max_col, max_row)

normalize_color(color) -> str
    "FF0000" → "FFFF0000"  # agrega opacidad si falta

create_fill(color) -> PatternFill
    Crea relleno sólido openpyxl.

create_border_side(style, color) -> Side
    Crea borde Side openpyxl.

apply_cell_style(cell, style_config) -> None
    Aplica estilo a una celda individual.

apply_style_to_range(ws, range_str, style_config) -> None
    Aplica estilo a un rango completo.

apply_borders_to_range(ws, range_str, color, style) -> None
    Bordes uniformes en un rango.

apply_borders_by_blocks(ws, range_str, row_step, col_step, ...) -> None
    Bordes en grilla por bloques.

apply_borders_to_sheet(ws, color, style) -> None
    Bordes en toda la hoja usada.

apply_cell_scaling(ws, scale_factor) -> None
    Escala filas y columnas.

merge_cell_ranges(ws, merge_ranges) -> None
    Fusiona rangos y centra contenido.

apply_conditional_formatting(ws, sheet_cfg) -> None
    Aplica reglas de formato condicional avanzadas.


═══════════════════════════════════════════════════════════════════════
EJEMPLOS
═══════════════════════════════════════════════════════════════════════
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.formatting.rule import FormulaRule
import re


# =============================================================================
# UTILIDADES DE PARSING
# =============================================================================

def parse_range(range_str: str) -> tuple:
    """Parsea un rango Excel y retorna (min_col, min_row, max_col, max_row)."""
    return range_boundaries(range_str)


def normalize_color(color: str) -> str:
    """Normaliza color a ARGB 8 chars. 'FF0000' → 'FFFF0000'."""
    color = str(color).replace("#", "").upper()
    if len(color) == 6:
        return f"FF{color}"
    return color


def create_border_side(style: str, color: str) -> Side:
    """Crea un Side de borde con estilo y color."""
    return Side(style=style, color=normalize_color(color))


def create_fill(color: str) -> PatternFill:
    """Crea un relleno sólido PatternFill."""
    normalized = normalize_color(color)
    return PatternFill(start_color=normalized, end_color=normalized, fill_type="solid")


# =============================================================================
# APLICACIÓN DE ESTILOS
# =============================================================================

def apply_cell_style(cell, style_config: dict) -> None:
    """
    Aplica estilos a una celda individual.

    style_config soporta:
        bold     : bool  — negrita
        color    : str   — color de fuente RGB
        align    : str   — 'center', 'left', 'right'
        bg_color : str   — color de fondo RGB
    """
    if 'bold' in style_config:
        font_color = style_config.get('color', '000000')
        cell.font = Font(bold=style_config['bold'], color=font_color)

    if 'align' in style_config:
        cell.alignment = Alignment(
            horizontal=style_config['align'],
            vertical='center'
        )

    if 'font_color' in style_config:
        normalized = normalize_color(style_config['font_color'])
        existing = cell.font
        cell.font = Font(bold=existing.bold, italic=existing.italic,
                         name=existing.name, size=existing.size, color=normalized)

    if 'bg_color' in style_config:
        cell.fill = create_fill(style_config['bg_color'])


def apply_style_to_range(ws, range_str: str, style_config: dict) -> None:
    """Aplica estilos a todas las celdas de un rango."""
    for row in ws[range_str]:
        for cell in row:
            apply_cell_style(cell, style_config)


# =============================================================================
# APLICACIÓN DE BORDES
# =============================================================================

def apply_borders_to_range(ws, range_str: str, color: str = "000000",
                           style: str = "thin") -> None:
    """Aplica bordes uniformes a un rango."""
    side = create_border_side(style, color)
    border = Border(left=side, right=side, top=side, bottom=side)

    for row in ws[range_str]:
        for cell in row:
            cell.border = border


def apply_borders_by_blocks(ws, range_str: str, row_step: int = 1,
                            col_step: int = 1, row_offset: int = 0,
                            col_offset: int = 0, color: str = "000000",
                            style: str = "thin") -> None:
    """
    Aplica bordes en grilla por bloques.

    Útil para tablas donde varias filas forman una unidad lógica
    (ej: turno = 2 filas, la par tiene nombre, la impar tiene aula).

    row_step / col_step: tamaño del bloque.
    row_offset / col_offset: celdas a saltar al inicio (skip headers).

    Las celdas internas del bloque reciben bordes finos; solo los
    límites del bloque reciben el borde principal.
    """
    if row_step <= 1 and col_step <= 1 and row_offset <= 0 and col_offset <= 0:
        apply_borders_to_range(ws, range_str, color, style)
        return

    side = create_border_side(style, color)
    thin_side = create_border_side("thin", color)

    min_col, min_row, max_col, max_row = parse_range(range_str)
    anchor_row = min_row + max(0, row_offset)
    anchor_col = min_col + max(0, col_offset)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if row < anchor_row or col < anchor_col:
                ws.cell(row=row, column=col).border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side
                )
                continue

            rel_row = row - anchor_row
            rel_col = col - anchor_col

            top = side if (rel_row % row_step == 0) else None
            left = side if (rel_col % col_step == 0) else None
            bottom = side if ((rel_row + 1) % row_step == 0 or row == max_row) else None
            right = side if ((rel_col + 1) % col_step == 0 or col == max_col) else None

            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom
            )


def apply_borders_to_sheet(ws, color: str = "000000", style: str = "thin") -> None:
    """Aplica bordes a todo el rango usado de la hoja."""
    if ws.max_row == 0 or ws.max_column == 0:
        return

    side = create_border_side(style, color)
    border = Border(left=side, right=side, top=side, bottom=side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


# =============================================================================
# FORMATO CONDICIONAL
# =============================================================================

def _replace_formula_placeholders(template: str, cell_ref: str, row: int,
                                   col_letter: str, next_cell: str = None) -> str:
    """
    Reemplaza placeholders en una plantilla de fórmula.

    Placeholders:
        {celda}          → "C4"
        {celda_siguiente} → "C5"
        {fila}           → "4"
        {columna}        → "C"
    """
    formula = template.replace("{celda}", cell_ref)
    formula = formula.replace("{fila}", str(row))
    formula = formula.replace("{columna}", col_letter)
    if next_cell:
        formula = formula.replace("{celda_siguiente}", next_cell)
    return formula


def _apply_conditional_rule(ws, cell_ref: str, formula: str,
                            fill: PatternFill, stop_if_true: bool) -> None:
    """Aplica una regla de formato condicional a una celda."""
    rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=stop_if_true)
    ws.conditional_formatting.add(cell_ref, rule)


def apply_conditional_formatting(ws, sheet_cfg: dict) -> None:
    """
    Aplica formato condicional avanzado con plantillas.

    Lee sheet_cfg['conditional_format_rules'].

    Tipos de regla:
        'rango'             → cada celda del rango
        'filas_pares'       → filas 1,3,5... relativas
        'filas_impares'     → filas 2,4,6... relativas
        'pares_con_siguiente' → evalúa par+siguiente, aplica a una
    """
    rules = sheet_cfg.get("conditional_format_rules", [])

    for rule_cfg in rules:
        rule_type = rule_cfg.get("tipo", "rango")
        range_str = rule_cfg.get("rango", "A1:A1")
        formula_template = rule_cfg.get("formula", "")
        color = rule_cfg.get("color", "FFFFFF")
        stop_if_true = rule_cfg.get("stop_if_true", True)
        row_step = max(1, int(rule_cfg.get("row_step", 2)))
        row_start_offset = int(rule_cfg.get("row_start_offset", 0))
        next_offset = max(0, int(rule_cfg.get("next_offset", 1)))

        fill = create_fill(color)
        min_col, min_row, max_col, max_row = parse_range(range_str)

        if rule_type == "rango":
            _apply_rule_to_all_cells(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true
            )

        elif rule_type == "filas_pares":
            _apply_rule_to_rows(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true,
                start_offset=row_start_offset, step=row_step
            )

        elif rule_type == "filas_impares":
            _apply_rule_to_rows(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true,
                start_offset=row_start_offset, step=row_step
            )

        elif rule_type == "pares_con_siguiente":
            target = rule_cfg.get("aplicar_a", "siguiente")
            _apply_rule_pairs_with_next(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true, target,
                start_offset=row_start_offset, step=row_step, next_offset=next_offset
            )


def _apply_rule_to_all_cells(ws, min_row, max_row, min_col, max_col,
                              formula_template, fill, stop_if_true):
    """Aplica regla a todas las celdas del rango."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            formula = _replace_formula_placeholders(
                formula_template, cell_ref, row, col_letter
            )
            _apply_conditional_rule(ws, cell_ref, formula, fill, stop_if_true)


def _apply_rule_to_rows(ws, min_row, max_row, min_col, max_col,
                         formula_template, fill, stop_if_true,
                         start_offset, step):
    """Aplica regla a filas específicas (pares o impares)."""
    for row_offset in range(start_offset, max_row - min_row + 1, step):
        actual_row = min_row + row_offset
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{actual_row}"
            formula = _replace_formula_placeholders(
                formula_template, cell_ref, actual_row, col_letter
            )
            _apply_conditional_rule(ws, cell_ref, formula, fill, stop_if_true)


def _apply_rule_pairs_with_next(ws, min_row, max_row, min_col, max_col,
                                 formula_template, fill, stop_if_true, target,
                                 start_offset=0, step=2, next_offset=1):
    """Aplica regla evaluando fila par y siguiente."""
    for row_offset in range(start_offset, max_row - min_row + 1, step):
        even_row = min_row + row_offset
        odd_row = even_row + next_offset
        if odd_row > max_row:
            break

        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            even_cell = f"{col_letter}{even_row}"
            odd_cell = f"{col_letter}{odd_row}"

            formula = _replace_formula_placeholders(
                formula_template, even_cell, even_row, col_letter, odd_cell
            )

            target_cell = odd_cell if target == "siguiente" else even_cell
            _apply_conditional_rule(ws, target_cell, formula, fill, stop_if_true)


# =============================================================================
# OTRAS UTILIDADES
# =============================================================================

def apply_cell_scaling(ws, scale_factor: float) -> None:
    """Escala alto de fila y ancho de columna por un factor."""
    BASE_ROW_HEIGHT = 15.0
    DEFAULT_COL_WIDTH = 8.43

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = BASE_ROW_HEIGHT * scale_factor

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[col_letter].width or DEFAULT_COL_WIDTH
        ws.column_dimensions[col_letter].width = current_width * scale_factor


def merge_cell_ranges(ws, merge_ranges: list) -> None:
    """Fusiona rangos y centra el contenido."""
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)
        top_left = ws[merge_range.split(":")[0]]
        top_left.alignment = Alignment(horizontal="center", vertical="center")


# =============================================================================
# PROCESAMIENTO DE HOJAS (etapas del pipeline)
# =============================================================================

def _process_headers(ws, sheet_cfg: dict) -> None:
    """Etapa 1: Encabezados."""
    headers = sheet_cfg.get("headers", [])
    if not headers or all(h == "" for h in headers):
        return

    ws.append(headers)
    header_style = sheet_cfg.get("header_style", {})
    for col in range(1, len(headers) + 1):
        apply_cell_style(ws.cell(row=1, column=col), header_style)


def _process_params(ws, sheet_cfg: dict) -> None:
    """Escribe valores de parámetros en celdas específicas antes de los datos."""
    for cell_ref, value in sheet_cfg.get("params", {}).items():
        ws[cell_ref] = value


def _process_data(ws, sheet_cfg: dict) -> None:
    """Etapa 2: Datos."""
    for row in sheet_cfg.get("data", []):
        ws.append(row)


def _process_formulas(ws, sheet_cfg: dict) -> None:
    """Etapa 3: Fórmulas simples por coordenada numérica."""
    for formula_def in sheet_cfg.get("formulas", []):
        ws.cell(
            row=formula_def["row"],
            column=formula_def["col"],
            value=formula_def["value"]
        )


def _process_column_widths(ws, sheet_cfg: dict) -> None:
    """Etapa 5: Anchos de columna con auto-ajuste al contenido más largo."""
    hidden = {i for i, w in sheet_cfg.get("column_widths", {}).items() if w == 0}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx in hidden:
            ws.column_dimensions[col_letter].hidden = True
            continue
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)


def _process_legacy_conditional_formatting(ws, sheet_cfg: dict) -> None:
    """Etapa 6: Formato condicional legacy."""
    for cf_def in sheet_cfg.get("conditional_formatting", []):
        fill = create_fill(cf_def["fill_color"])
        rule = FormulaRule(formula=[cf_def["rule"]], fill=fill, stopIfTrue=True)
        ws.conditional_formatting.add(cf_def["range"], rule)


def _process_cell_scaling(ws, sheet_cfg: dict) -> None:
    """Etapa 7: Escalado de celdas."""
    cell_size = sheet_cfg.get("cell_size")
    if cell_size is not None:
        apply_cell_scaling(ws, float(cell_size))


def _process_table_borders(ws, sheet_cfg: dict) -> None:
    """Etapa 8: Bordes (uniformes o por bloques)."""
    if not sheet_cfg.get("table_borders", False):
        return

    border_color = sheet_cfg.get("border_color", "000000")
    border_style = sheet_cfg.get("border_style", "thin")
    border_step = int(sheet_cfg.get("border_step", 1))

    # Mapeo de tamaños de bloque por rango
    block_sizes = {}
    for item in sheet_cfg.get("table_block_sizes", []):
        if isinstance(item, dict) and item.get("range"):
            block_sizes[item["range"]] = (
                int(item.get("row_step", border_step)),
                int(item.get("col_step", border_step)),
                1 if item.get("skip_first_row") else 0,
                1 if item.get("skip_first_col") else 0,
            )

    table_ranges = sheet_cfg.get("table_ranges", [])

    if block_sizes:
        for block_range, (row_step, col_step, row_offset, col_offset) in block_sizes.items():
            apply_borders_by_blocks(
                ws, block_range,
                row_step=row_step, col_step=col_step,
                row_offset=row_offset, col_offset=col_offset,
                color=border_color, style=border_style
            )
    elif table_ranges:
        for table_range in table_ranges:
            apply_borders_by_blocks(
                ws, table_range,
                row_step=border_step, col_step=border_step,
                color=border_color, style=border_style
            )
    else:
        if border_step > 1:
            used_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            apply_borders_by_blocks(
                ws, used_range,
                row_step=border_step, col_step=border_step,
                color=border_color, style=border_style
            )
        else:
            apply_borders_to_sheet(ws, color=border_color, style=border_style)


def _process_range_styles(ws, sheet_cfg: dict) -> None:
    """Etapa 9: Estilos por rango (colores de fondo/fuente)."""
    for style_def in sheet_cfg.get("range_styles", []):
        target_range = style_def.get("range")
        style = style_def.get("style", {})
        if target_range and style:
            apply_style_to_range(ws, target_range, style)


def _process_merge_ranges(ws, sheet_cfg: dict) -> None:
    """Etapa 10: Celdas fusionadas."""
    merge_ranges = sheet_cfg.get("merge_ranges", [])
    if merge_ranges:
        merge_cell_ranges(ws, merge_ranges)


def _process_conditional_format_rules(ws, sheet_cfg: dict) -> None:
    """Etapa 11: Formato condicional avanzado."""
    if sheet_cfg.get("conditional_format_rules"):
        apply_conditional_formatting(ws, sheet_cfg)


def _process_conditional_formats(ws, sheet_cfg: dict) -> None:
    """Etapa 12: FormulaRule SUMPRODUCT generados por el DSL (modo :excel-cf).

    Cada entrada de 'conditional_formats' tiene:
        range   : str  — rango Excel al que se aplica la regla (ej. "B4:F6")
        formula : str  — fórmula SUMPRODUCT (sin '=' inicial)
        style   : dict — puede tener font_color y/o bg_color
    """
    for cf_def in sheet_cfg.get("conditional_formats", []):
        range_str = cf_def.get("range")
        formula   = cf_def.get("formula", "")
        style     = cf_def.get("style", {})
        if not range_str or not formula:
            continue

        font = None
        if "font_color" in style:
            font = Font(color=normalize_color(style["font_color"]), bold=True)

        fill = None
        if "bg_color" in style:
            fill = create_fill(style["bg_color"])

        rule = FormulaRule(formula=[formula], font=font, fill=fill, stopIfTrue=False)
        ws.conditional_formatting.add(range_str, rule)


def _inject_fernando_formulas(ws, sheet_cfg: dict) -> None:
    """
    Etapa 4: Inyecta fórmulas complejas con normalización automática.

    Lee sheet_cfg['fernando_formulas'].

    Normalizaciones aplicadas:
        1. [$C111.$C$9]     → C111!$C$9   (referencias ODS)
        2. $C111.$D5         → C111!$D5    (punto → exclamación)
        3. c112!$d5          → C112!$D5    (mayúsculas)
        4. CONCATENAR(a;b;c) → (a & b & c) (compatibilidad)
        5. SUSTITUIR         → SUBSTITUTE  (localización ES→EN)
        6. ESPACIOS          → TRIM
        7. SI                → IF
        8. ;                 → ,           (separador)

    Segundo pase defensivo: re-normaliza cualquier fórmula que ya
    exista en la hoja por si fue escrita en los datos.
    """
    def _normalize_fernando_formula(formula: str) -> str:
        f = formula

        def _split_top_level_args(arg_str: str) -> list[str]:
            """Separa argumentos de nivel superior respetando paréntesis y comillas."""
            args = []
            current = []
            depth = 0
            in_quotes = False
            i = 0
            while i < len(arg_str):
                ch = arg_str[i]
                if ch == '"':
                    in_quotes = not in_quotes
                    current.append(ch)
                elif not in_quotes and ch == '(':
                    depth += 1
                    current.append(ch)
                elif not in_quotes and ch == ')':
                    depth -= 1
                    current.append(ch)
                elif not in_quotes and depth == 0 and ch in (';', ','):
                    args.append(''.join(current).strip())
                    current = []
                else:
                    current.append(ch)
                i += 1
            tail = ''.join(current).strip()
            if tail:
                args.append(tail)
            return args

        def _rewrite_concat_calls(expr: str) -> str:
            """Reescribe CONCAT/CONCATENAR a concatenación con &."""
            fn_pattern = re.compile(r'(?i)\b(COM\.MICROSOFT\.CONCAT|CONCATENAR|CONCATENATE|CONCAT)\(')
            while True:
                match = fn_pattern.search(expr)
                if not match:
                    break

                open_paren = match.end() - 1
                depth = 0
                in_quotes = False
                close_paren = None
                for idx in range(open_paren, len(expr)):
                    ch = expr[idx]
                    if ch == '"':
                        in_quotes = not in_quotes
                    elif not in_quotes and ch == '(':
                        depth += 1
                    elif not in_quotes and ch == ')':
                        depth -= 1
                        if depth == 0:
                            close_paren = idx
                            break

                if close_paren is None:
                    break

                inner = expr[open_paren + 1:close_paren]
                args = _split_top_level_args(inner)
                if len(args) <= 1:
                    replacement = f"({inner})"
                else:
                    replacement = "(" + " & ".join(args) + ")"
                expr = expr[:match.start()] + replacement + expr[close_paren + 1:]
            return expr

        # 1) Referencias ODS con corchetes
        f = re.sub(r'\[\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)\$(\d+)\]', r'\1!$\2$\3', f)
        f = re.sub(r'\[\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)(\d+)\]', r'\1!$\2\3', f)
        f = re.sub(r'\[\$([A-Za-z0-9_]+)\.([A-Za-z]+)(\d+)\]', r'\1!\2\3', f)

        # 2) Referencias con punto
        f = re.sub(r'\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)\$(\d+)', r'\1!$\2$\3', f)
        f = re.sub(r'\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)(\d+)', r'\1!$\2\3', f)
        f = re.sub(r'\$([A-Za-z0-9_]+)\.([A-Za-z]+)(\d+)', r'\1!\2\3', f)

        # 3) Mayúsculas en nombres de hoja/columna
        def _upper_sheet_col(match):
            sheet, col, row = match.group(1), match.group(2), match.group(3)
            return f"{sheet.upper()}!${col.upper()}{row}"

        def _upper_sheet_col_row(match):
            sheet, col, row = match.group(1), match.group(2), match.group(3)
            return f"{sheet.upper()}!${col.upper()}${row}"

        f = re.sub(r'([A-Za-z0-9_]+)!\$([A-Za-z]+)(\d+)', _upper_sheet_col, f)
        f = re.sub(r'([A-Za-z0-9_]+)!\$([A-Za-z]+)\$(\d+)', _upper_sheet_col_row, f)

        # 3.5) Reescribir CONCAT a &
        f = _rewrite_concat_calls(f)

        # 4) Traducir funciones ES → EN
        localized_map = {
            'SUSTITUIR': 'SUBSTITUTE',
            'ESPACIOS': 'TRIM',
            'CONCATENAR': 'CONCATENATE',
            'SI': 'IF',
            'CONTAR.SI': 'COUNTIF',
        }
        for local_func, excel_func in sorted(localized_map.items(), key=lambda x: len(x[0]), reverse=True):
            pattern = rf'(?<![A-Z0-9_.]){re.escape(local_func)}(?=\()'
            f = re.sub(pattern, excel_func, f, flags=re.IGNORECASE)

        f = f.replace(';', ',')

        return f

    formulas = sheet_cfg.get('fernando_formulas', [])
    count = 0
    for formula_def in formulas:
        cell_ref = formula_def.get('cell')
        formula = formula_def.get('formula') or formula_def.get('excel')

        if cell_ref and formula:
            try:
                cell = ws[cell_ref]
                cell.value = _normalize_fernando_formula(formula)
                count += 1
            except Exception:
                pass

    # Segundo pase defensivo
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                normalized = _normalize_fernando_formula(value)
                if normalized != value:
                    cell.value = normalized

    if count > 0:
        print(f"   ✅ Inyectadas {count} fórmulas en hoja '{ws.title}'")


# =============================================================================
# FUNCIÓN PRINCIPAL
# =============================================================================

def _offset_range_rows(range_str: str, offset: int) -> str:
    """Desplaza los números de fila en un rango tipo 'A1:K7' por offset."""
    import re
    def shift(cell):
        m = re.match(r'([A-Z]+)(\d+)', cell)
        return f"{m.group(1)}{int(m.group(2)) + offset}" if m else cell
    return ":".join(shift(p) for p in range_str.split(":"))


def _unwrap_regions(sheet_cfg: dict) -> dict:
    """
    Convierte formato regions a formato plano (backward compat).
    Las hojas generadas por el DSL Lisp vienen con 'regions';
    las hojas escritas a mano en Python vienen planas.

    Una región: copia sus claves a la raíz.
    Múltiples regiones: apila verticalmente con 1 fila en blanco entre ellas.
    """
    if "regions" not in sheet_cfg:
        return sheet_cfg
    regions = sheet_cfg["regions"]
    if not regions:
        return sheet_cfg
    if len(regions) == 1:
        for k, v in regions[0].items():
            if k not in sheet_cfg:
                sheet_cfg[k] = v
        return sheet_cfg

    GAP_ROWS = 1
    all_data = []
    all_formulas = []
    all_table_ranges = []
    all_block_sizes = []
    current_row = 1  # fila Excel 1-indexed donde empieza la región actual

    for i, region in enumerate(regions):
        if i > 0:
            ncols = len(all_data[-1]) if all_data else 0
            all_data.append([""] * ncols)
            current_row += GAP_ROWS

        region_data = region.get("data", [])
        row_offset = current_row - 1

        for row in region_data:
            all_data.append(row)

        for tr in region.get("table_ranges", []):
            all_table_ranges.append(_offset_range_rows(tr, row_offset))

        for bs in region.get("table_block_sizes", []):
            adjusted = dict(bs)
            adjusted["range"] = _offset_range_rows(bs["range"], row_offset)
            all_block_sizes.append(adjusted)

        for f in region.get("formulas", []):
            all_formulas.append({**f, "row": f["row"] + row_offset})

        current_row += len(region_data)

    first = regions[0]
    result = {k: v for k, v in sheet_cfg.items() if k != "regions"}
    result["data"] = all_data
    result["table_ranges"] = all_table_ranges
    result["table_block_sizes"] = all_block_sizes
    if all_formulas:
        result["formulas"] = all_formulas
    for k in ["headers", "column_widths", "border_color", "border_style"]:
        if k in first and k not in result:
            result[k] = first[k]
    return result


def _apply_defaults(sheet_cfg: dict) -> dict:
    """
    Aplica valores por defecto para estilo cuando el DSL no los proporciona.
    El DSL ahora es puramente estructural; el estilo es responsabilidad del backend.
    """
    sheet_cfg.setdefault("table_borders", True)
    sheet_cfg.setdefault("border_color", "000000")
    sheet_cfg.setdefault("border_style", "thin")
    sheet_cfg.setdefault("header_style", {})
    hs = sheet_cfg["header_style"]
    if "bold" not in hs:
        hs["bold"] = True
    if "align" not in hs:
        hs["align"] = "center"
    if "bg_color" not in hs:
        hs["bg_color"] = "4A90E2"
    return sheet_cfg


def generate_excel(config: dict, filename: str) -> None:
    """
    Genera un archivo .xlsx completo desde una configuración declarativa.

    Args:
        config: Diccionario con clave "sheets" (lista de configuraciones).
        filename: Ruta del archivo de salida.

    Ver docstring del módulo para referencia completa de configuración.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_cfg in config.get("sheets", []):
        sheet_cfg = _unwrap_regions(sheet_cfg)
        sheet_cfg = _apply_defaults(sheet_cfg)
        ws = wb.create_sheet(title=sheet_cfg["title"])

        # Pipeline de procesamiento (orden importante)
        _process_headers(ws, sheet_cfg)
        _process_data(ws, sheet_cfg)
        _process_params(ws, sheet_cfg)
        _process_formulas(ws, sheet_cfg)
        _inject_fernando_formulas(ws, sheet_cfg)
        _process_column_widths(ws, sheet_cfg)
        _process_legacy_conditional_formatting(ws, sheet_cfg)
        _process_cell_scaling(ws, sheet_cfg)
        _process_table_borders(ws, sheet_cfg)
        _process_range_styles(ws, sheet_cfg)
        _process_merge_ranges(ws, sheet_cfg)
        _process_conditional_format_rules(ws, sheet_cfg)
        _process_conditional_formats(ws, sheet_cfg)

    wb.save(filename)
    print(f"Archivo '{filename}' generado exitosamente.")


# Alias para compatibilidad con código existente
generar_excel_personalizado = generate_excel


# =============================================================================
# API PARAMETRICA (SIN JSON MANUAL DESDE LISP)
# =============================================================================

DEFAULT_DAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
DEFAULT_AULAS_CATALOGO = [f"Aula {i}" for i in range(1, 10)] + ["Lab"]


def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _build_row_names(turnos: int, row_step: int) -> list:
    row_names = []
    for turno in range(1, max(1, turnos) + 1):
        row_names.append(f"Turno {turno}")
        row_names.extend([""] * max(0, row_step - 1))
        if turno == 3 and turno < turnos:
            row_names.append("")
    return row_names


def _build_dynamic_merge_ranges(turnos: int, row_step: int) -> list:
    ranges = []
    for turno in range(1, max(1, turnos) + 1):
        offset = 1 if (turno >= 4 and turnos >= 4) else 0
        start_row = 4 + ((turno - 1) * row_step) + offset
        end_row = start_row + max(0, row_step - 1)
        ranges.append(f"B{start_row}:B{end_row}")
    return ranges


def _normalize_subject_row(item) -> list:
    if isinstance(item, dict):
        abrev = item.get("abrev", item.get("abreviatura", ""))
        nombre = item.get("asignatura", item.get("nombre", ""))
        frec = item.get("frec", item.get("frecuencia", ""))
        faltan = item.get("faltan", "")
        asignadas = item.get("asignadas", "")
        return [abrev, nombre, frec, faltan, asignadas]

    if isinstance(item, (list, tuple)):
        values = list(item[:5])
        while len(values) < 5:
            values.append("")
        return values

    raise ValueError(
        "Formato de asignatura invalido. Usa dict o tuple/list con al menos (abrev, nombre, frec)."
    )


def _normalize_horario_rows(horario_data, expected_rows: int, day_count: int) -> list:
    source = list(horario_data or [])
    normalized = []

    for i in range(expected_rows):
        row = source[i] if i < len(source) else []
        row_values = list(row) if isinstance(row, (list, tuple)) else []
        row_values = row_values[:day_count] + [""] * max(0, day_count - len(row_values))
        normalized.append([_safe_text(v) for v in row_values])

    return normalized


def _extract_unique_aulas(horario_rows: list, row_step: int) -> list:
    seen = set()
    aulas = []
    offset = 1 if row_step > 1 else 0

    for row_idx in range(offset, len(horario_rows), max(1, row_step)):
        row = horario_rows[row_idx]
        for raw_value in row:
            value = _safe_text(raw_value).strip()
            if value and value not in seen:
                seen.add(value)
                aulas.append(value)

    return aulas


def construir_hoja_grupo_desde_parametros(
    grupo: str,
    horario_data,
    asignaturas_data=None,
    turnos: int = 6,
    row_step: int = 3,
    days=None,
    aulas_catalogo=None,
    column_width: float = 14,
    border_color: str = "4F81BD",
    border_style: str = "medium",
) -> dict:
    days = list(days or DEFAULT_DAYS)
    aulas_catalogo = list(aulas_catalogo or DEFAULT_AULAS_CATALOGO)

    day_start_col = 3
    day_end_col = day_start_col + len(days) - 1
    day_end_col_letter = get_column_letter(day_end_col)

    row_names = _build_row_names(turnos, row_step)
    total_horario_rows = len(row_names)
    horario_end_row = 3 + total_horario_rows

    horario_rows = _normalize_horario_rows(horario_data, total_horario_rows, len(days))
    normalized_subjects = [_normalize_subject_row(s) for s in (asignaturas_data or [])]

    asig_height = max(1, len(normalized_subjects))
    asig_end_row = 3 + asig_height
    dynamic_asig_range = f"I4:M{asig_end_row}"
    asig_abrev_range = f"$I4:I{asig_end_row}"

    aulas_detectadas = _extract_unique_aulas(horario_rows, row_step)
    if not aulas_detectadas:
        aulas_detectadas = aulas_catalogo
    aulas_lateral = aulas_detectadas + [""] * max(0, total_horario_rows - len(aulas_detectadas))
    aulas_lateral = aulas_lateral[:total_horario_rows]

    aulas_end_row = 3 + max(1, len([a for a in aulas_lateral if _safe_text(a).strip()]))
    dynamic_aulas_range = f"O4:O{aulas_end_row}"
    aulas_range_cf = f"$O4:O{aulas_end_row}"

    has_separator = "Turno 4" in row_names
    horario_ranges = []
    turnos_ranges = []
    separator_range = None

    if has_separator:
        expected_turno4_index = 3 * row_step
        actual_turno4_index = row_names.index("Turno 4")
        pre_end_row = 3 + expected_turno4_index
        post_start_row = 4 + actual_turno4_index
        separator_row = 4 + expected_turno4_index

        turnos_ranges = [f"B4:B{pre_end_row}", f"B{post_start_row}:B{horario_end_row}"]
        horario_ranges = [
            f"C4:{day_end_col_letter}{pre_end_row}",
            f"C{post_start_row}:{day_end_col_letter}{horario_end_row}",
        ]
        separator_range = f"B{separator_row}:{day_end_col_letter}{separator_row}"
    else:
        turnos_ranges = [f"B4:B{horario_end_row}"]
        horario_ranges = [f"C4:{day_end_col_letter}{horario_end_row}"]

    data = [
        ["Grupo ", grupo] + [""] * 13,
        [""] * 15,
        ["", "", *days, "", "Abrev", "Asignaturas", "Frec", "Faltan", "Asignadas", "", "Aulas"],
    ]

    for idx in range(total_horario_rows):
        asig_row = normalized_subjects[idx] if idx < len(normalized_subjects) else ["", "", "", "", ""]
        data.append([
            "",
            row_names[idx],
            *horario_rows[idx],
            "",
            *asig_row,
            "",
            _safe_text(aulas_lateral[idx]),
        ])

    formulas = [
        {"row": asig_end_row + 1, "col": 12, "value": "Total:"},
        {"row": asig_end_row + 1, "col": 13, "value": f"=COUNTA(I4:I{asig_end_row})"},
        {"row": asig_end_row + 2, "col": 12, "value": "Σ Frec:"},
        {"row": asig_end_row + 2, "col": 13, "value": f"=SUM(K4:K{asig_end_row})"},
        {"row": aulas_end_row + 1, "col": 14, "value": "Total:"},
        {"row": aulas_end_row + 1, "col": 15, "value": f"=COUNTA(O4:O{aulas_end_row})"},
        {"row": horario_end_row + 1, "col": 6, "value": "Ocupados:"},
        {
            "row": horario_end_row + 1,
            "col": 7,
            "value": f"=COUNTA(C4:{day_end_col_letter}{horario_end_row})/{max(1, row_step)}",
        },
    ]

    for idx, asig in enumerate(normalized_subjects, start=4):
        abrev = _safe_text(asig[0]).strip()
        if not abrev:
            continue
        formulas.append({
            "row": idx,
            "col": 13,
            "value": f"=COUNTIF(C4:{day_end_col_letter}{horario_end_row},I{idx})",
        })
        formulas.append({"row": idx, "col": 12, "value": f"=K{idx}-M{idx}"})

    conditional_format_rules = []
    for horario_range in horario_ranges:
        conditional_format_rules.extend([
            {
                "tipo": "filas_pares",
                "rango": horario_range,
                "formula": f'AND({{celda}}<>"", COUNTIF({asig_abrev_range},{{celda}})=0)',
                "color": "F4A460",
                "row_step": row_step,
                "row_start_offset": 0,
            },
            {
                "tipo": "filas_impares",
                "rango": horario_range,
                "formula": f'AND({{celda}}<>"", COUNTIF({aulas_range_cf},{{celda}})=0)',
                "color": "FFD700",
                "row_step": row_step,
                "row_start_offset": 1,
            },
            {
                "tipo": "pares_con_siguiente",
                "rango": horario_range,
                "formula": 'AND({celda}<>"", {celda_siguiente}="")',
                "color": "FF0000",
                "row_step": row_step,
                "next_offset": 1 if row_step > 1 else 0,
                "aplicar_a": "siguiente",
            },
        ])

    conditional_format_rules.extend([
        {
            "tipo": "rango",
            "rango": f"J4:J{asig_end_row}",
            "formula": 'AND({celda}<>"", M{fila}>0, L{fila}=0)',
            "color": "00FF00",
        },
        {
            "tipo": "rango",
            "rango": f"J4:J{asig_end_row}",
            "formula": 'AND({celda}<>"", M{fila}>0, L{fila}<0)',
            "color": "FF6B6B",
        },
        {
            "tipo": "rango",
            "rango": f"J4:J{asig_end_row}",
            "formula": 'AND({celda}<>"", M{fila}>0, L{fila}>0, L{fila}<K{fila})',
            "color": "FFA500",
        },
    ])

    table_ranges = ["B3:" + day_end_col_letter + "3", *turnos_ranges]
    if separator_range:
        table_ranges.append(separator_range)
    table_ranges.extend([*horario_ranges, "I3:M3", dynamic_asig_range, dynamic_aulas_range])

    table_block_sizes = [
        {"range": f"B3:{day_end_col_letter}3", "row_step": 1, "col_step": 1},
        *[{"range": r, "row_step": row_step, "col_step": 1} for r in turnos_ranges],
        *[{"range": r, "row_step": row_step, "col_step": 1} for r in horario_ranges],
        {"range": "I3:M3", "row_step": 1, "col_step": 1},
        {"range": dynamic_asig_range, "row_step": 1, "col_step": 1},
        {"range": dynamic_aulas_range, "row_step": 1, "col_step": 1},
    ]
    if separator_range:
        table_block_sizes.insert(1 + len(turnos_ranges), {"range": separator_range, "row_step": 1, "col_step": 1})

    range_styles = [
        {"range": f"I3:I{asig_end_row}", "style": {"bg_color": "A9D18E"}},
        *[{"range": r, "style": {"bg_color": "F4CCCC"}} for r in turnos_ranges],
    ]

    return {
        "title": grupo,
        "data": data,
        "column_widths": {i: column_width for i in range(1, 16)},
        "range_styles": range_styles,
        "table_ranges": table_ranges,
        "horario_data_range": f"C4:{day_end_col_letter}{horario_end_row}",
        "table_block_sizes": table_block_sizes,
        "merge_ranges": _build_dynamic_merge_ranges(turnos, row_step),
        "table_borders": True,
        "border_color": border_color,
        "border_style": border_style,
        "formulas": formulas,
        "conditional_format_rules": conditional_format_rules,
    }


def _build_turno_labels(turnos: int) -> list:
    base = ["1ro", "2do", "3ro", "4to", "5to", "6to"]
    if turnos <= len(base):
        return base[:turnos]
    return base + [f"Turno {i}" for i in range(7, turnos + 1)]


def _build_day_blocks(turnos: int, day_count: int) -> list:
    blocks = []
    row_offset = 3
    for day_idx in range(day_count):
        header_row = row_offset
        row_start = header_row + 1
        row_end = row_start + turnos - 1
        group_col_letter = get_column_letter(3 + day_idx)
        blocks.append((header_row, row_start, row_end, group_col_letter))
        row_offset = row_end + (2 if day_idx < day_count - 1 else 0)
    return blocks


def _build_aulas_formula(groups: list, group_cell_ref: str, header_ref: str) -> str:
    parts = [
        f'IF({group}!{group_cell_ref}={header_ref},{group}!$B$1 & " ","")'
        for group in groups
    ]
    return f'=SUBSTITUTE(TRIM(CONCAT({",".join(parts)}))," ",",")'


def _build_aulas_fernando_formulas(
    groups: list,
    turnos: int,
    row_step: int,
    day_count: int,
    aulas_count: int,
) -> list:
    if not groups:
        return []

    formulas = []
    aula_offset = 1 if row_step > 1 else 0

    for header_row, row_start, row_end, group_col in _build_day_blocks(turnos, day_count):
        for aulas_col in range(3, 3 + aulas_count):
            aulas_col_letter = get_column_letter(aulas_col)
            header_ref = f"{aulas_col_letter}${header_row}"

            for aulas_row in range(row_start, row_end + 1):
                turno_index = aulas_row - row_start
                turno_offset = 1 if turno_index >= 3 and turnos >= 4 else 0
                group_row = 4 + aula_offset + (turno_index * row_step) + turno_offset
                group_cell_ref = f"${group_col}${group_row}"
                cell_ref = f"{aulas_col_letter}{aulas_row}"

                formulas.append({
                    "cell": cell_ref,
                    "formula": _build_aulas_formula(groups, group_cell_ref, header_ref),
                })

    return formulas


def construir_hoja_aulas_desde_parametros(
    grupos: list,
    aulas_por_dia=None,
    turnos: int = 6,
    row_step: int = 3,
    days=None,
    aulas_catalogo=None,
    column_width: float = 12,
    border_color: str = "B3B3B3",
    border_style: str = "thick",
) -> dict:
    days = list(days or DEFAULT_DAYS)
    aulas_catalogo = list(aulas_catalogo or DEFAULT_AULAS_CATALOGO)
    aulas_por_dia = aulas_por_dia or {}

    aulas_header_labels = []
    for idx, aula_name in enumerate(aulas_catalogo, start=1):
        name = _safe_text(aula_name)
        if name.lower().startswith("aula "):
            aulas_header_labels.append(name.split(" ", 1)[1])
        elif name.lower() == "laboratorio":
            aulas_header_labels.append("Lab")
        else:
            aulas_header_labels.append(name if name else str(idx))

    turno_labels = _build_turno_labels(turnos)
    data = [[""] * (2 + len(aulas_catalogo)), ["", "", *aulas_header_labels]]

    for idx, day in enumerate(days):
        day_rows = list(aulas_por_dia.get(day, []))
        data.append(["", day, *aulas_catalogo])

        for turno_idx in range(turnos):
            row = day_rows[turno_idx] if turno_idx < len(day_rows) else []
            row_values = list(row) if isinstance(row, (list, tuple)) else []
            row_values = row_values[:len(aulas_catalogo)] + [""] * max(0, len(aulas_catalogo) - len(row_values))
            data.append(["", turno_labels[turno_idx], *[_safe_text(v) for v in row_values]])

        if idx < len(days) - 1:
            data.append([""] * (2 + len(aulas_catalogo)))

    end_col_letter = get_column_letter(2 + len(aulas_catalogo))
    table_ranges = [
        f"B{header_row}:{end_col_letter}{row_end}"
        for header_row, _row_start, row_end, _group_col in _build_day_blocks(turnos, len(days))
    ]
    range_styles = [
        {"range": f"B{header_row}:{end_col_letter}{header_row}", "style": {"bg_color": "E6B8AF"}}
        for header_row, _row_start, _row_end, _group_col in _build_day_blocks(turnos, len(days))
    ]

    return {
        "title": "Aulas",
        "data": data,
        "column_widths": {i: column_width for i in range(1, 3 + len(aulas_catalogo))},
        "range_styles": range_styles,
        "table_ranges": table_ranges,
        "table_borders": True,
        "border_color": border_color,
        "border_style": border_style,
        "fernando_formulas": _build_aulas_fernando_formulas(
            groups=list(grupos),
            turnos=turnos,
            row_step=row_step,
            day_count=len(days),
            aulas_count=len(aulas_catalogo),
        ),
    }


def construir_config_desde_parametros(
    grupos,
    horarios_por_grupo,
    asignaturas_por_grupo=None,
    aulas_por_dia=None,
    turnos: int = 6,
    horario_row_step: int = 3,
    days=None,
    aulas_catalogo=None,
) -> dict:
    grupos = list(grupos or [])
    horarios_por_grupo = horarios_por_grupo or {}
    asignaturas_por_grupo = asignaturas_por_grupo or {}

    if not grupos:
        grupos = list(horarios_por_grupo.keys())

    sheets = []
    for grupo in grupos:
        sheets.append(construir_hoja_grupo_desde_parametros(
            grupo=grupo,
            horario_data=horarios_por_grupo.get(grupo, []),
            asignaturas_data=asignaturas_por_grupo.get(grupo, []),
            turnos=turnos,
            row_step=horario_row_step,
            days=days,
            aulas_catalogo=aulas_catalogo,
        ))

    sheets.append(construir_hoja_aulas_desde_parametros(
        grupos=grupos,
        aulas_por_dia=aulas_por_dia,
        turnos=turnos,
        row_step=horario_row_step,
        days=days,
        aulas_catalogo=aulas_catalogo,
    ))

    return {"sheets": sheets}


def generar_excel_desde_parametros(
    filename: str,
    grupos,
    horarios_por_grupo,
    asignaturas_por_grupo=None,
    aulas_por_dia=None,
    turnos: int = 6,
    horario_row_step: int = 3,
    days=None,
    aulas_catalogo=None,
) -> None:
    config = construir_config_desde_parametros(
        grupos=grupos,
        horarios_por_grupo=horarios_por_grupo,
        asignaturas_por_grupo=asignaturas_por_grupo,
        aulas_por_dia=aulas_por_dia,
        turnos=turnos,
        horario_row_step=horario_row_step,
        days=days,
        aulas_catalogo=aulas_catalogo,
    )
    generate_excel(config, filename)


generate_excel_from_parameters = generar_excel_desde_parametros
build_config_from_parameters = construir_config_desde_parametros


def _normalizar_dia_tv(dia: str) -> str:
    value = _safe_text(dia).strip().lower()
    mapping = {
        "lunes": "Lunes",
        "martes": "Martes",
        "miercoles": "Miércoles",
        "miércoles": "Miércoles",
        "jueves": "Jueves",
        "viernes": "Viernes",
        "sabado": "Sábado",
        "sábado": "Sábado",
        "domingo": "Domingo",
    }
    if value in mapping:
        return mapping[value]
    return _safe_text(dia).strip().title() or "Dia"


def _sheet_title_safe(title: str, fallback: str = "Hoja") -> str:
    value = _safe_text(title).strip() or fallback
    for ch in ('[', ']', ':', '*', '?', '/', '\\'):
        value = value.replace(ch, "-")
    return value[:31] or fallback


def _extraer_programas_unicos_tv(programas: list) -> list:
    seen = set()
    result = []
    for item in programas:
        name = _safe_text(item.get("nombre", "")).strip()
        key = name.lower()
        if not name or key in seen:
            continue
        seen.add(key)
        result.append({
            "nombre": name,
            "duracion": item.get("duracion", 0),
            "tipo_programa": _safe_text(item.get("tipo_programa", "")),
        })
    return result


def construir_hoja_tv_desde_parametros(
    dia_cfg: dict,
    index: int = 1,
    start_time_param_cell: str = "I2",
) -> dict:
    dia_display = _normalizar_dia_tv(dia_cfg.get("dia", f"dia-{index}"))
    title = _sheet_title_safe(f"{index:02d}-{dia_display}", f"Dia-{index}")
    planificacion = list(dia_cfg.get("programas", []) or [])
    programas = _extraer_programas_unicos_tv(planificacion)

    max_rows = max(1, len(programas), len(planificacion))
    default_start = _safe_text(planificacion[0].get("hora_inicio", "00:00")) if planificacion else "00:00"
    prog_end_row = 3 + max(1, len(programas))
    plan_end_row = 3 + max(1, len(planificacion))

    data = []
    data.append([f"Programacion {dia_display}", "", "", "", "", "", "", "", ""])
    data.append(["", "", "", "", "", "", "", "Hora inicio 1ra fila", default_start])
    data.append(["Programa", "Duracion (min)", "Tipo", "", "Hora de inicio", "Hora de terminacion", "Programa", "Tipo calc", ""])

    for idx in range(max_rows):
        prog = programas[idx] if idx < len(programas) else {}
        plan = planificacion[idx] if idx < len(planificacion) else {}
        data.append([
            _safe_text(prog.get("nombre", "")),
            prog.get("duracion", ""),
            _safe_text(prog.get("tipo_programa", "")),
            "",
            "",
            "",
            _safe_text(plan.get("nombre", "")),
            "",
            "",
        ])

    formulas = []
    for idx in range(len(planificacion)):
        row = 4 + idx
        if idx == 0:
            formulas.append({
                "row": row,
                "col": 5,
                "value": f'=IF(G{row}="","",${start_time_param_cell.replace("$", "")})',
            })
        else:
            formulas.append({
                "row": row,
                "col": 5,
                "value": f'=IF(G{row}="","",F{row - 1})',
            })

        formulas.append({
            "row": row,
            "col": 6,
            "value": f'=IF(G{row}="","",TEXT(TIMEVALUE(E{row}) + IFERROR(VLOOKUP(G{row},$A$4:$C${prog_end_row},2,FALSE),0)/1440,"hh:mm"))',
        })

        formulas.append({
            "row": row,
            "col": 8,
            "value": f'=IF(G{row}="","",IFERROR(VLOOKUP(G{row},$A$4:$C${prog_end_row},3,FALSE),""))',
        })

    conditional_format_rules = []
    if len(planificacion) >= 1:
        plan_range = f"E4:G{plan_end_row}"
        conditional_format_rules.extend([
            {
                "tipo": "rango",
                "rango": plan_range,
                "formula": 'AND($H{fila}<>"", $H{fila}=IFERROR(INDEX($H:$H,{fila}-1),""), $H{fila}=IFERROR(INDEX($H:$H,{fila}+1),""))',
                "color": "FF6B6B",
            },
            {
                "tipo": "rango",
                "rango": plan_range,
                "formula": 'AND($H{fila}<>"", $H{fila}=IFERROR(INDEX($H:$H,{fila}-1),""), NOT($H{fila}=IFERROR(INDEX($H:$H,{fila}+1),"")))',
                "color": "FFA500",
            },
            {
                "tipo": "rango",
                "rango": plan_range,
                "formula": 'AND($H{fila}<>"", $H{fila}=IFERROR(INDEX($H:$H,{fila}+1),""), NOT($H{fila}=IFERROR(INDEX($H:$H,{fila}-1),"")))',
                "color": "FFA500",
            },
        ])

    return {
        "title": title,
        "data": data,
        "column_widths": {1: 42, 2: 14, 3: 16, 4: 3, 5: 16, 6: 18, 7: 42, 8: 12, 9: 18},
        "table_borders": True,
        "border_color": "B7B7B7",
        "border_style": "thin",
        "table_ranges": [f"A3:C{prog_end_row}", f"E3:G{plan_end_row}", "H2:I2"],
        "range_styles": [
            {"range": "A3:C3", "style": {"bold": True, "align": "center", "bg_color": "D9EAD3"}},
            {"range": "E3:G3", "style": {"bold": True, "align": "center", "bg_color": "CFE2F3"}},
            {"range": "H2:I2", "style": {"bold": True, "align": "center", "bg_color": "FFF2CC"}},
        ],
        "formulas": formulas,
        "conditional_format_rules": conditional_format_rules,
    }


def construir_hoja_resumen_tv(nombre_canal: str, planificacion_semanal: list) -> dict:
    data = []
    total_programas = 0
    total_minutos = 0

    for dia_cfg in planificacion_semanal:
        programas = dia_cfg.get("programas", []) or []
        dia_display = _normalizar_dia_tv(dia_cfg.get("dia", ""))
        cantidad = len(programas)
        total_programas += cantidad

        minutos = 0
        tipos = set()
        publicos = set()
        for p in programas:
            try:
                minutos += int(p.get("duracion", 0) or 0)
            except (TypeError, ValueError):
                pass
            tipo = _safe_text(p.get("tipo_programa", "")).strip()
            publico = _safe_text(p.get("tipo_publico", "")).strip()
            if tipo:
                tipos.add(tipo)
            if publico:
                publicos.add(publico)

        total_minutos += minutos
        inicio = _safe_text(programas[0].get("hora_inicio", "")) if programas else ""
        fin = _safe_text(programas[-1].get("hora_final", "")) if programas else ""

        data.append([
            dia_display,
            cantidad,
            minutos,
            inicio,
            fin,
            ", ".join(sorted(tipos)),
            ", ".join(sorted(publicos)),
        ])

    data.append(["TOTAL", total_programas, total_minutos, "", "", "", _safe_text(nombre_canal)])

    max_row = max(1, len(data) + 1)
    return {
        "title": _sheet_title_safe("Resumen TV", "Resumen"),
        "headers": ["Día", "Programas", "Minutos", "Inicio", "Fin", "Tipos", "Públicos"],
        "data": data,
        "column_widths": {1: 14, 2: 12, 3: 12, 4: 10, 5: 10, 6: 36, 7: 28},
        "header_style": {"bold": True, "align": "center", "bg_color": "F4CCCC"},
        "table_borders": True,
        "border_color": "B7B7B7",
        "border_style": "thin",
        "table_ranges": [f"A1:G{max_row}"],
        "range_styles": [{"range": f"A{max_row}:G{max_row}", "style": {"bold": True, "bg_color": "FFF2CC"}}],
    }


def construir_config_horario_tv_desde_parametros(
    nombre_canal: str,
    planificacion_semanal,
    incluir_resumen: bool = True,
) -> dict:
    plan = list(planificacion_semanal or [])
    sheets = [
        construir_hoja_tv_desde_parametros(day_cfg, index=idx)
        for idx, day_cfg in enumerate(plan, start=1)
    ]

    if incluir_resumen:
        sheets.append(construir_hoja_resumen_tv(nombre_canal, plan))

    return {"sheets": sheets}


def generar_excel_horario_tv_desde_parametros(
    filename: str,
    nombre_canal: str,
    planificacion_semanal,
    incluir_resumen: bool = True,
) -> None:
    config = construir_config_horario_tv_desde_parametros(
        nombre_canal=nombre_canal,
        planificacion_semanal=planificacion_semanal,
        incluir_resumen=incluir_resumen,
    )
    generate_excel(config, filename)


build_tv_schedule_config_from_parameters = construir_config_horario_tv_desde_parametros
generate_tv_schedule_excel_from_parameters = generar_excel_horario_tv_desde_parametros


# =============================================================================
# EJEMPLOS DE USO
# =============================================================================

if __name__ == "__main__":

    # ── Ejemplo 1: Tabla simple ──────────────────────────────────────
    example_config = {
        "sheets": [
            {
                "title": "Ejemplo",
                "headers": ["ID", "Nombre", "Valor"],
                "data": [
                    [1, "Item A", 100],
                    [2, "Item B", 200],
                    [3, "Item C", 300],
                ],
                "column_widths": {1: 10, 2: 20, 3: 15},
                "header_style": {"bold": True, "align": "center"},
                "table_borders": True,
                "border_color": "000000",
                "border_style": "thin",
            }
        ]
    }

    generate_excel(example_config, "ejemplo_refactorizado.xlsx")
