"""
hoja_con_formulas.py — Framework de Generación de Excel
========================================================

Genera archivos .xlsx completos con datos, fórmulas, estilos, bordes,
formato condicional y celdas fusionadas a partir de una configuración
declarativa en Python (diccionarios).

Puede ser alimentado por cualquier generador externo: Common Lisp, JSON,
YAML, bases de datos, etc.

Dependencias
------------
    pip install openpyxl

Uso mínimo
----------
    from hoja_con_formulas import generate_excel

    generate_excel({
        "sheets": [{
            "title": "Mi Hoja",
            "data": [["A", "B"], [1, 2]],
            "table_borders": True
        }]
    }, "salida.xlsx")

Alias de compatibilidad
-----------------------
    generar_excel_personalizado = generate_excel


═══════════════════════════════════════════════════════════════════════
ARQUITECTURA: Pipeline de procesamiento
═══════════════════════════════════════════════════════════════════════

generate_excel() recorre cada hoja y aplica estas etapas en orden:

    1. _process_headers()          → Encabezados (opcional, usa ws.append)
    2. _process_data()             → Filas de datos (ws.append)
    3. _process_formulas()         → Fórmulas simples por coordenada
    4. _inject_fernando_formulas() → Fórmulas complejas (cross-sheet)
    5. _process_column_widths()    → Ancho de columnas
    6. _process_legacy_conditional_formatting() → Formato condicional legacy
    7. _process_cell_scaling()     → Escalar filas/columnas por factor
    8. _process_table_borders()    → Bordes (uniformes o por bloques)
    9. _process_range_styles()     → Colores de fondo / fuente por rango
   10. _process_merge_ranges()     → Celdas fusionadas
   11. _process_conditional_format_rules() → Formato condicional avanzado

IMPORTANTE: Los estilos de rango (paso 9) se aplican DESPUÉS de los datos
(paso 2), por lo que sobrescriben cualquier formato previo. El orden es
deliberado: datos → fórmulas → bordes → colores → fusiones → condicional.


═══════════════════════════════════════════════════════════════════════
REFERENCIA DE CONFIGURACIÓN DE HOJA
═══════════════════════════════════════════════════════════════════════

Cada elemento de config["sheets"] acepta las siguientes claves:

───────────────────────────────────────────────────────────────────────
OBLIGATORIAS
───────────────────────────────────────────────────────────────────────

title : str
    Nombre de la hoja en el libro Excel.

───────────────────────────────────────────────────────────────────────
DATOS
───────────────────────────────────────────────────────────────────────

data : list[list]
    Filas de datos. Cada sublista es una fila.
    Se agregan con ws.append(), por lo que la primera sublista va en la
    fila 1 (o fila 2 si se usó "headers").

    Ejemplo:
        "data": [
            ["", "", "1", "2", "3"],
            ["", "Lunes", "Aula 1", "Aula 2", "Aula 3"],
            ["", "1ro", "", "", ""],
        ]

headers : list[str]  (opcional)
    Fila de encabezados. Se inserta ANTES de data con ws.append().
    Se estiliza con header_style.

    Ejemplo:
        "headers": ["ID", "Nombre", "Valor"]

───────────────────────────────────────────────────────────────────────
FÓRMULAS SIMPLES
───────────────────────────────────────────────────────────────────────

formulas : list[dict]  (opcional)
    Fórmulas o valores en celdas específicas, referenciadas por
    coordenadas numéricas (row, col) 1-based.

    Cada dict tiene:
        row   : int   — número de fila (1-based)
        col   : int   — número de columna (1-based)
        value : str   — fórmula (empieza con "=") o valor literal

    Ejemplo:
        "formulas": [
            {"row": 5, "col": 13, "value": "=COUNTIF($C$4:$G$15,I5)"},
            {"row": 5, "col": 12, "value": "=K5-M5"},
            {"row": 11, "col": 12, "value": "Total:"},
        ]

───────────────────────────────────────────────────────────────────────
FÓRMULAS COMPLEJAS (Cross-Sheet)
───────────────────────────────────────────────────────────────────────

fernando_formulas : list[dict]  (opcional)
    Fórmulas complejas referenciadas por notación Excel (A1, C4, etc.).
    Diseñado para fórmulas que cruzan múltiples hojas.

    Cada dict tiene:
        cell    : str — referencia Excel ("C4")
        formula : str — la fórmula (puede estar en español o inglés)
        # o alternativamente:
        excel   : str — la fórmula (se acepta "formula" o "excel")

    El motor normaliza automáticamente:
        • Referencias ODS [$C111.$C$9] → Excel C111!$C$9
        • Referencias con punto $C111.$D5 → C111!$D5
        • Funciones en español SI() → IF()
        • Funciones en español SUSTITUIR() → SUBSTITUTE()
        • Funciones en español ESPACIOS() → TRIM()
        • CONCATENAR(a;b;c) → (a & b & c)
        • Separadores ; → ,

    Ejemplo:
        "fernando_formulas": [
            {
                "cell": "C4",
                "formula": "=SUBSTITUTE(TRIM(CONCAT("
                    "IF(C111!$C$5=C$3,C111!$B$1 & \" \", \"\"),"
                    "IF(C121!$C$5=C$3,C121!$B$1 & \" \", \"\"))), "
                    "\" \", \",\")"
            }
        ]

───────────────────────────────────────────────────────────────────────
COLUMNAS
───────────────────────────────────────────────────────────────────────

column_widths : dict[int, float]  (opcional)
    Ancho de columnas por índice (1-based).

    Ejemplo:
        "column_widths": {1: 14, 2: 14, 3: 20}

        # O con comprensión:
        "column_widths": {i: 12 for i in range(1, 13)}

cell_size : float  (opcional)
    Factor de escala para TODAS las celdas de la hoja.
    2.0 = doble tamaño. Afecta alto de fila y ancho de columna.

    Ejemplo:
        "cell_size": 1.5

───────────────────────────────────────────────────────────────────────
ESTILOS DE ENCABEZADO
───────────────────────────────────────────────────────────────────────

header_style : dict  (opcional)
    Estilo aplicado a la fila de "headers".

    Claves soportadas:
        bold     : bool   — negrita
        color    : str    — color de fuente (RGB 6 chars)
        align    : str    — 'center', 'left', 'right'
        bg_color : str    — color de fondo (RGB 6 chars)

    Ejemplo:
        "header_style": {"bold": True, "align": "center", "bg_color": "E6B8AF"}

───────────────────────────────────────────────────────────────────────
ESTILOS POR RANGO
───────────────────────────────────────────────────────────────────────

range_styles : list[dict]  (opcional)
    Aplica estilos a rangos específicos de celdas.

    Cada dict tiene:
        range : str  — rango Excel ("B3:L3")
        style : dict — mismas claves que header_style

    Ejemplo:
        "range_styles": [
            {"range": "B3:L3",  "style": {"bg_color": "E6B8AF"}},
            {"range": "B11:L11", "style": {"bg_color": "E6B8AF"}},
            {"range": "I3:I10", "style": {"bg_color": "A9D18E"}},
        ]

───────────────────────────────────────────────────────────────────────
BORDES
───────────────────────────────────────────────────────────────────────

table_borders : bool  (opcional)
    Activa el sistema de bordes.

border_color : str  (opcional, default "000000")
    Color del borde en hex RGB.

border_style : str  (opcional, default "thin")
    Estilo del borde: "thin", "medium", "thick".

border_step : int  (opcional, default 1)
    Tamaño base de bloque para bordes agrupados.

table_ranges : list[str]  (opcional)
    Lista de rangos Excel donde aplicar bordes.
    Si no se especifica, se aplica a toda la hoja usada.

    Ejemplo:
        "table_ranges": ["B3:L9", "B11:L17", "B19:L25"]

table_block_sizes : list[dict]  (opcional)
    Configura bordes por bloques para rangos específicos.
    Útil cuando las celdas están agrupadas (ej: turnos de 2 filas).

    Cada dict tiene:
        range         : str   — rango Excel
        row_step      : int   — altura del bloque en filas
        col_step      : int   — ancho del bloque en columnas
        skip_first_row : bool — saltar primera fila del bloque
        skip_first_col : bool — saltar primera columna del bloque

    Ejemplo (turnos de 2 filas):
        "table_block_sizes": [
            {"range": "B4:B15", "row_step": 2, "col_step": 1},
            {"range": "C4:G15", "row_step": 2, "col_step": 1},
        ]

    Sin table_block_sizes, los bordes se aplican celda por celda.

───────────────────────────────────────────────────────────────────────
CELDA FUSIONADAS
───────────────────────────────────────────────────────────────────────

merge_ranges : list[str]  (opcional)
    Lista de rangos a fusionar. El contenido se centra automáticamente.

    Ejemplo:
        "merge_ranges": ["B4:B5", "B6:B7", "B8:B9"]

───────────────────────────────────────────────────────────────────────
FORMATO CONDICIONAL (AVANZADO)
───────────────────────────────────────────────────────────────────────

conditional_format_rules : list[dict]  (opcional)
    Reglas de formato condicional con plantillas de fórmula.

    Cada dict tiene:
        tipo    : str  — tipo de aplicación (ver abajo)
        rango   : str  — rango Excel donde aplicar
        formula : str  — plantilla de fórmula con placeholders
        color   : str  — color de relleno cuando se cumple
        stop_if_true : bool  — detener evaluación si cumple (default True)

        # Solo para tipo "pares_con_siguiente":
        aplicar_a : str  — "siguiente" o "actual"

    Placeholders en la fórmula:
        {celda}          — referencia actual (ej: "C4")
        {celda_siguiente} — referencia fila siguiente (ej: "C5")
        {fila}           — número de fila (ej: "4")
        {columna}        — letra de columna (ej: "C")

    Tipos de regla:
        "rango"          — aplica a cada celda del rango
        "filas_pares"    — aplica a filas 1, 3, 5... relativas al rango
        "filas_impares"  — aplica a filas 2, 4, 6... relativas al rango
        "pares_con_siguiente" — evalúa par+siguiente, aplica a una

    Ejemplo:
        "conditional_format_rules": [
            {
                "tipo": "filas_pares",
                "rango": "C4:G15",
                "formula": 'AND({celda}<>"", COUNTIF($I4:I10,{celda})=0)',
                "color": "F4A460"
            },
            {
                "tipo": "pares_con_siguiente",
                "rango": "C4:G15",
                "formula": 'AND({celda}<>"", {celda_siguiente}="")',
                "color": "FF0000",
                "aplicar_a": "siguiente"
            },
        ]

───────────────────────────────────────────────────────────────────────
FORMATO CONDICIONAL (LEGACY)
───────────────────────────────────────────────────────────────────────

conditional_formatting : list[dict]  (opcional)
    Formato condicional simple (formato legacy).

    Cada dict tiene:
        range     : str  — rango Excel
        rule      : str  — fórmula (sin placeholders)
        fill_color: str  — color de relleno

    Ejemplo:
        "conditional_formatting": [
            {
                "range": "A1:A10",
                "rule": 'A1>100',
                "fill_color": "FF0000"
            }
        ]


═══════════════════════════════════════════════════════════════════════
API PÚBLICA
═══════════════════════════════════════════════════════════════════════

generate_excel(config: dict, filename: str) -> None
    Función principal. Genera el archivo Excel.

generar_excel_personalizado(config: dict, filename: str) -> None
    Alias de generate_excel para compatibilidad.

───────────────────────────────────────────────────────────────────────
Funciones reutilizables (uso directo opcional)
───────────────────────────────────────────────────────────────────────

parse_range(range_str) -> tuple
    Parsea "C4:G17" → (3, 4, 7, 17)  # (min_col, min_row, max_col, max_row)

normalize_color(color) -> str
    "FF0000" → "FFFF0000"  # agrega opacidad si falta

create_fill(color) -> PatternFill
    Crea relleno sólido openpyxl.

create_border_side(style, color) -> Side
    Crea borde Side openpyxl.

apply_cell_style(cell, style_config) -> None
    Aplica estilo a una celda individual.

apply_style_to_range(ws, range_str, style_config) -> None
    Aplica estilo a un rango completo.

apply_borders_to_range(ws, range_str, color, style) -> None
    Bordes uniformes en un rango.

apply_borders_by_blocks(ws, range_str, row_step, col_step, ...) -> None
    Bordes en grilla por bloques.

apply_borders_to_sheet(ws, color, style) -> None
    Bordes en toda la hoja usada.

apply_cell_scaling(ws, scale_factor) -> None
    Escala filas y columnas.

merge_cell_ranges(ws, merge_ranges) -> None
    Fusiona rangos y centra contenido.

apply_conditional_formatting(ws, sheet_cfg) -> None
    Aplica reglas de formato condicional avanzadas.


═══════════════════════════════════════════════════════════════════════
EJEMPLOS
═══════════════════════════════════════════════════════════════════════
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.formatting.rule import FormulaRule
import re


# =============================================================================
# UTILIDADES DE PARSING
# =============================================================================

def parse_range(range_str: str) -> tuple:
    """Parsea un rango Excel y retorna (min_col, min_row, max_col, max_row)."""
    return range_boundaries(range_str)


def normalize_color(color: str) -> str:
    """Normaliza color a ARGB 8 chars. 'FF0000' → 'FFFF0000'."""
    color = str(color).replace("#", "").upper()
    if len(color) == 6:
        return f"FF{color}"
    return color


def create_border_side(style: str, color: str) -> Side:
    """Crea un Side de borde con estilo y color."""
    return Side(style=style, color=normalize_color(color))


def create_fill(color: str) -> PatternFill:
    """Crea un relleno sólido PatternFill."""
    normalized = normalize_color(color)
    return PatternFill(start_color=normalized, end_color=normalized, fill_type="solid")


# =============================================================================
# APLICACIÓN DE ESTILOS
# =============================================================================

def apply_cell_style(cell, style_config: dict) -> None:
    """
    Aplica estilos a una celda individual.

    style_config soporta:
        bold     : bool  — negrita
        color    : str   — color de fuente RGB
        align    : str   — 'center', 'left', 'right'
        bg_color : str   — color de fondo RGB
    """
    if 'bold' in style_config:
        font_color = style_config.get('color', '000000')
        cell.font = Font(bold=style_config['bold'], color=font_color)

    if 'align' in style_config:
        cell.alignment = Alignment(
            horizontal=style_config['align'],
            vertical='center'
        )

    if 'bg_color' in style_config:
        cell.fill = create_fill(style_config['bg_color'])


def apply_style_to_range(ws, range_str: str, style_config: dict) -> None:
    """Aplica estilos a todas las celdas de un rango."""
    for row in ws[range_str]:
        for cell in row:
            apply_cell_style(cell, style_config)


# =============================================================================
# APLICACIÓN DE BORDES
# =============================================================================

def apply_borders_to_range(ws, range_str: str, color: str = "000000",
                           style: str = "thin") -> None:
    """Aplica bordes uniformes a un rango."""
    side = create_border_side(style, color)
    border = Border(left=side, right=side, top=side, bottom=side)

    for row in ws[range_str]:
        for cell in row:
            cell.border = border


def apply_borders_by_blocks(ws, range_str: str, row_step: int = 1,
                            col_step: int = 1, row_offset: int = 0,
                            col_offset: int = 0, color: str = "000000",
                            style: str = "thin") -> None:
    """
    Aplica bordes en grilla por bloques.

    Útil para tablas donde varias filas forman una unidad lógica
    (ej: turno = 2 filas, la par tiene nombre, la impar tiene aula).

    row_step / col_step: tamaño del bloque.
    row_offset / col_offset: celdas a saltar al inicio (skip headers).

    Las celdas internas del bloque reciben bordes finos; solo los
    límites del bloque reciben el borde principal.
    """
    if row_step <= 1 and col_step <= 1 and row_offset <= 0 and col_offset <= 0:
        apply_borders_to_range(ws, range_str, color, style)
        return

    side = create_border_side(style, color)
    thin_side = create_border_side("thin", color)

    min_col, min_row, max_col, max_row = parse_range(range_str)
    anchor_row = min_row + max(0, row_offset)
    anchor_col = min_col + max(0, col_offset)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if row < anchor_row or col < anchor_col:
                ws.cell(row=row, column=col).border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side
                )
                continue

            rel_row = row - anchor_row
            rel_col = col - anchor_col

            top = side if (rel_row % row_step == 0) else None
            left = side if (rel_col % col_step == 0) else None
            bottom = side if ((rel_row + 1) % row_step == 0 or row == max_row) else None
            right = side if ((rel_col + 1) % col_step == 0 or col == max_col) else None

            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom
            )


def apply_borders_to_sheet(ws, color: str = "000000", style: str = "thin") -> None:
    """Aplica bordes a todo el rango usado de la hoja."""
    if ws.max_row == 0 or ws.max_column == 0:
        return

    side = create_border_side(style, color)
    border = Border(left=side, right=side, top=side, bottom=side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


# =============================================================================
# FORMATO CONDICIONAL
# =============================================================================

def _replace_formula_placeholders(template: str, cell_ref: str, row: int,
                                   col_letter: str, next_cell: str = None) -> str:
    """
    Reemplaza placeholders en una plantilla de fórmula.

    Placeholders:
        {celda}          → "C4"
        {celda_siguiente} → "C5"
        {fila}           → "4"
        {columna}        → "C"
    """
    formula = template.replace("{celda}", cell_ref)
    formula = formula.replace("{fila}", str(row))
    formula = formula.replace("{columna}", col_letter)
    if next_cell:
        formula = formula.replace("{celda_siguiente}", next_cell)
    return formula


def _apply_conditional_rule(ws, cell_ref: str, formula: str,
                            fill: PatternFill, stop_if_true: bool) -> None:
    """Aplica una regla de formato condicional a una celda."""
    rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=stop_if_true)
    ws.conditional_formatting.add(cell_ref, rule)


def apply_conditional_formatting(ws, sheet_cfg: dict) -> None:
    """
    Aplica formato condicional avanzado con plantillas.

    Lee sheet_cfg['conditional_format_rules'].

    Tipos de regla:
        'rango'             → cada celda del rango
        'filas_pares'       → filas 1,3,5... relativas
        'filas_impares'     → filas 2,4,6... relativas
        'pares_con_siguiente' → evalúa par+siguiente, aplica a una
    """
    rules = sheet_cfg.get("conditional_format_rules", [])

    for rule_cfg in rules:
        rule_type = rule_cfg.get("tipo", "rango")
        range_str = rule_cfg.get("rango", "A1:A1")
        formula_template = rule_cfg.get("formula", "")
        color = rule_cfg.get("color", "FFFFFF")
        stop_if_true = rule_cfg.get("stop_if_true", True)
        row_step = max(1, int(rule_cfg.get("row_step", 2)))
        row_start_offset = int(rule_cfg.get("row_start_offset", 0))
        next_offset = max(0, int(rule_cfg.get("next_offset", 1)))

        fill = create_fill(color)
        min_col, min_row, max_col, max_row = parse_range(range_str)

        if rule_type == "rango":
            _apply_rule_to_all_cells(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true
            )

        elif rule_type == "filas_pares":
            _apply_rule_to_rows(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true,
                start_offset=row_start_offset, step=row_step
            )

        elif rule_type == "filas_impares":
            _apply_rule_to_rows(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true,
                start_offset=row_start_offset, step=row_step
            )

        elif rule_type == "pares_con_siguiente":
            target = rule_cfg.get("aplicar_a", "siguiente")
            _apply_rule_pairs_with_next(
                ws, min_row, max_row, min_col, max_col,
                formula_template, fill, stop_if_true, target,
                start_offset=row_start_offset, step=row_step, next_offset=next_offset
            )


def _apply_rule_to_all_cells(ws, min_row, max_row, min_col, max_col,
                              formula_template, fill, stop_if_true):
    """Aplica regla a todas las celdas del rango."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            formula = _replace_formula_placeholders(
                formula_template, cell_ref, row, col_letter
            )
            _apply_conditional_rule(ws, cell_ref, formula, fill, stop_if_true)


def _apply_rule_to_rows(ws, min_row, max_row, min_col, max_col,
                         formula_template, fill, stop_if_true,
                         start_offset, step):
    """Aplica regla a filas específicas (pares o impares)."""
    for row_offset in range(start_offset, max_row - min_row + 1, step):
        actual_row = min_row + row_offset
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{actual_row}"
            formula = _replace_formula_placeholders(
                formula_template, cell_ref, actual_row, col_letter
            )
            _apply_conditional_rule(ws, cell_ref, formula, fill, stop_if_true)


def _apply_rule_pairs_with_next(ws, min_row, max_row, min_col, max_col,
                                 formula_template, fill, stop_if_true, target,
                                 start_offset=0, step=2, next_offset=1):
    """Aplica regla evaluando fila par y siguiente."""
    for row_offset in range(start_offset, max_row - min_row + 1, step):
        even_row = min_row + row_offset
        odd_row = even_row + next_offset
        if odd_row > max_row:
            break

        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            even_cell = f"{col_letter}{even_row}"
            odd_cell = f"{col_letter}{odd_row}"

            formula = _replace_formula_placeholders(
                formula_template, even_cell, even_row, col_letter, odd_cell
            )

            target_cell = odd_cell if target == "siguiente" else even_cell
            _apply_conditional_rule(ws, target_cell, formula, fill, stop_if_true)


# =============================================================================
# OTRAS UTILIDADES
# =============================================================================

def apply_cell_scaling(ws, scale_factor: float) -> None:
    """Escala alto de fila y ancho de columna por un factor."""
    BASE_ROW_HEIGHT = 15.0
    DEFAULT_COL_WIDTH = 8.43

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = BASE_ROW_HEIGHT * scale_factor

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[col_letter].width or DEFAULT_COL_WIDTH
        ws.column_dimensions[col_letter].width = current_width * scale_factor


def merge_cell_ranges(ws, merge_ranges: list) -> None:
    """Fusiona rangos y centra el contenido."""
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)
        top_left = ws[merge_range.split(":")[0]]
        top_left.alignment = Alignment(horizontal="center", vertical="center")


# =============================================================================
# PROCESAMIENTO DE HOJAS (etapas del pipeline)
# =============================================================================

def _process_headers(ws, sheet_cfg: dict) -> None:
    """Etapa 1: Encabezados."""
    headers = sheet_cfg.get("headers", [])
    if not headers:
        return

    ws.append(headers)
    header_style = sheet_cfg.get("header_style", {})
    for col in range(1, len(headers) + 1):
        apply_cell_style(ws.cell(row=1, column=col), header_style)


def _process_data(ws, sheet_cfg: dict) -> None:
    """Etapa 2: Datos."""
    for row in sheet_cfg.get("data", []):
        ws.append(row)


def _process_formulas(ws, sheet_cfg: dict) -> None:
    """Etapa 3: Fórmulas simples por coordenada numérica."""
    for formula_def in sheet_cfg.get("formulas", []):
        ws.cell(
            row=formula_def["row"],
            column=formula_def["col"],
            value=formula_def["value"]
        )


def _process_column_widths(ws, sheet_cfg: dict) -> None:
    """Etapa 5: Anchos de columna."""
    for col_idx, width in sheet_cfg.get("column_widths", {}).items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _process_legacy_conditional_formatting(ws, sheet_cfg: dict) -> None:
    """Etapa 6: Formato condicional legacy."""
    for cf_def in sheet_cfg.get("conditional_formatting", []):
        fill = create_fill(cf_def["fill_color"])
        rule = FormulaRule(formula=[cf_def["rule"]], fill=fill, stopIfTrue=True)
        ws.conditional_formatting.add(cf_def["range"], rule)


def _process_cell_scaling(ws, sheet_cfg: dict) -> None:
    """Etapa 7: Escalado de celdas."""
    cell_size = sheet_cfg.get("cell_size")
    if cell_size is not None:
        apply_cell_scaling(ws, float(cell_size))


def _process_table_borders(ws, sheet_cfg: dict) -> None:
    """Etapa 8: Bordes (uniformes o por bloques)."""
    if not sheet_cfg.get("table_borders", False):
        return

    border_color = sheet_cfg.get("border_color", "000000")
    border_style = sheet_cfg.get("border_style", "thin")
    border_step = int(sheet_cfg.get("border_step", 1))

    # Mapeo de tamaños de bloque por rango
    block_sizes = {}
    for item in sheet_cfg.get("table_block_sizes", []):
        if isinstance(item, dict) and item.get("range"):
            block_sizes[item["range"]] = (
                int(item.get("row_step", border_step)),
                int(item.get("col_step", border_step)),
                1 if item.get("skip_first_row") else 0,
                1 if item.get("skip_first_col") else 0,
            )

    table_ranges = sheet_cfg.get("table_ranges", [])

    if table_ranges:
        for table_range in table_ranges:
            row_step, col_step, row_offset, col_offset = block_sizes.get(
                table_range, (border_step, border_step, 0, 0)
            )
            apply_borders_by_blocks(
                ws, table_range,
                row_step=row_step, col_step=col_step,
                row_offset=row_offset, col_offset=col_offset,
                color=border_color, style=border_style
            )
    else:
        if border_step > 1:
            used_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            apply_borders_by_blocks(
                ws, used_range,
                row_step=border_step, col_step=border_step,
                color=border_color, style=border_style
            )
        else:
            apply_borders_to_sheet(ws, color=border_color, style=border_style)


def _process_range_styles(ws, sheet_cfg: dict) -> None:
    """Etapa 9: Estilos por rango (colores de fondo/fuente)."""
    for style_def in sheet_cfg.get("range_styles", []):
        target_range = style_def.get("range")
        style = style_def.get("style", {})
        if target_range and style:
            apply_style_to_range(ws, target_range, style)


def _process_merge_ranges(ws, sheet_cfg: dict) -> None:
    """Etapa 10: Celdas fusionadas."""
    merge_ranges = sheet_cfg.get("merge_ranges", [])
    if merge_ranges:
        merge_cell_ranges(ws, merge_ranges)


def _process_conditional_format_rules(ws, sheet_cfg: dict) -> None:
    """Etapa 11: Formato condicional avanzado."""
    if sheet_cfg.get("conditional_format_rules"):
        apply_conditional_formatting(ws, sheet_cfg)


def _inject_fernando_formulas(ws, sheet_cfg: dict) -> None:
    """
    Etapa 4: Inyecta fórmulas complejas con normalización automática.

    Lee sheet_cfg['fernando_formulas'].

    Normalizaciones aplicadas:
        1. [$C111.$C$9]     → C111!$C$9   (referencias ODS)
        2. $C111.$D5         → C111!$D5    (punto → exclamación)
        3. c112!$d5          → C112!$D5    (mayúsculas)
        4. CONCATENAR(a;b;c) → (a & b & c) (compatibilidad)
        5. SUSTITUIR         → SUBSTITUTE  (localización ES→EN)
        6. ESPACIOS          → TRIM
        7. SI                → IF
        8. ;                 → ,           (separador)

    Segundo pase defensivo: re-normaliza cualquier fórmula que ya
    exista en la hoja por si fue escrita en los datos.
    """
    def _normalize_fernando_formula(formula: str) -> str:
        f = formula

        def _split_top_level_args(arg_str: str) -> list[str]:
            """Separa argumentos de nivel superior respetando paréntesis y comillas."""
            args = []
            current = []
            depth = 0
            in_quotes = False
            i = 0
            while i < len(arg_str):
                ch = arg_str[i]
                if ch == '"':
                    in_quotes = not in_quotes
                    current.append(ch)
                elif not in_quotes and ch == '(':
                    depth += 1
                    current.append(ch)
                elif not in_quotes and ch == ')':
                    depth -= 1
                    current.append(ch)
                elif not in_quotes and depth == 0 and ch in (';', ','):
                    args.append(''.join(current).strip())
                    current = []
                else:
                    current.append(ch)
                i += 1
            tail = ''.join(current).strip()
            if tail:
                args.append(tail)
            return args

        def _rewrite_concat_calls(expr: str) -> str:
            """Reescribe CONCAT/CONCATENAR a concatenación con &."""
            fn_pattern = re.compile(r'(?i)\b(COM\.MICROSOFT\.CONCAT|CONCATENAR|CONCATENATE|CONCAT)\(')
            while True:
                match = fn_pattern.search(expr)
                if not match:
                    break

                open_paren = match.end() - 1
                depth = 0
                in_quotes = False
                close_paren = None
                for idx in range(open_paren, len(expr)):
                    ch = expr[idx]
                    if ch == '"':
                        in_quotes = not in_quotes
                    elif not in_quotes and ch == '(':
                        depth += 1
                    elif not in_quotes and ch == ')':
                        depth -= 1
                        if depth == 0:
                            close_paren = idx
                            break

                if close_paren is None:
                    break

                inner = expr[open_paren + 1:close_paren]
                args = _split_top_level_args(inner)
                if len(args) <= 1:
                    replacement = f"({inner})"
                else:
                    replacement = "(" + " & ".join(args) + ")"
                expr = expr[:match.start()] + replacement + expr[close_paren + 1:]
            return expr

        # 1) Referencias ODS con corchetes
        f = re.sub(r'\[\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)\$(\d+)\]', r'\1!$\2$\3', f)
        f = re.sub(r'\[\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)(\d+)\]', r'\1!$\2\3', f)
        f = re.sub(r'\[\$([A-Za-z0-9_]+)\.([A-Za-z]+)(\d+)\]', r'\1!\2\3', f)

        # 2) Referencias con punto
        f = re.sub(r'\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)\$(\d+)', r'\1!$\2$\3', f)
        f = re.sub(r'\$([A-Za-z0-9_]+)\.\$([A-Za-z]+)(\d+)', r'\1!$\2\3', f)
        f = re.sub(r'\$([A-Za-z0-9_]+)\.([A-Za-z]+)(\d+)', r'\1!\2\3', f)

        # 3) Mayúsculas en nombres de hoja/columna
        def _upper_sheet_col(match):
            sheet, col, row = match.group(1), match.group(2), match.group(3)
            return f"{sheet.upper()}!${col.upper()}{row}"

        def _upper_sheet_col_row(match):
            sheet, col, row = match.group(1), match.group(2), match.group(3)
            return f"{sheet.upper()}!${col.upper()}${row}"

        f = re.sub(r'([A-Za-z0-9_]+)!\$([A-Za-z]+)(\d+)', _upper_sheet_col, f)
        f = re.sub(r'([A-Za-z0-9_]+)!\$([A-Za-z]+)\$(\d+)', _upper_sheet_col_row, f)

        # 3.5) Reescribir CONCAT a &
        f = _rewrite_concat_calls(f)

        # 4) Traducir funciones ES → EN
        localized_map = {
            'SUSTITUIR': 'SUBSTITUTE',
            'ESPACIOS': 'TRIM',
            'CONCATENAR': 'CONCATENATE',
            'SI': 'IF',
            'CONTAR.SI': 'COUNTIF',
        }
        for local_func, excel_func in sorted(localized_map.items(), key=lambda x: len(x[0]), reverse=True):
            pattern = rf'(?<![A-Z0-9_.]){re.escape(local_func)}(?=\()'
            f = re.sub(pattern, excel_func, f, flags=re.IGNORECASE)

        f = f.replace(';', ',')

        return f

    formulas = sheet_cfg.get('fernando_formulas', [])
    count = 0
    for formula_def in formulas:
        cell_ref = formula_def.get('cell')
        formula = formula_def.get('formula') or formula_def.get('excel')

        if cell_ref and formula:
            try:
                cell = ws[cell_ref]
                cell.value = _normalize_fernando_formula(formula)
                count += 1
            except Exception:
                pass

    # Segundo pase defensivo
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                normalized = _normalize_fernando_formula(value)
                if normalized != value:
                    cell.value = normalized

    if count > 0:
        print(f"   ✅ Inyectadas {count} fórmulas en hoja '{ws.title}'")


# =============================================================================
# FUNCIÓN PRINCIPAL
# =============================================================================

def generate_excel(config: dict, filename: str) -> None:
    """
    Genera un archivo .xlsx completo desde una configuración declarativa.

    Args:
        config: Diccionario con clave "sheets" (lista de configuraciones).
        filename: Ruta del archivo de salida.

    Ver docstring del módulo para referencia completa de configuración.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_cfg in config.get("sheets", []):
        ws = wb.create_sheet(title=sheet_cfg["title"])

        # Pipeline de procesamiento (orden importante)
        _process_headers(ws, sheet_cfg)
        _process_data(ws, sheet_cfg)
        _process_formulas(ws, sheet_cfg)
        _inject_fernando_formulas(ws, sheet_cfg)
        _process_column_widths(ws, sheet_cfg)
        _process_legacy_conditional_formatting(ws, sheet_cfg)
        _process_cell_scaling(ws, sheet_cfg)
        _process_table_borders(ws, sheet_cfg)
        _process_range_styles(ws, sheet_cfg)
        _process_merge_ranges(ws, sheet_cfg)
        _process_conditional_format_rules(ws, sheet_cfg)

    wb.save(filename)
    print(f"Archivo '{filename}' generado exitosamente.")


# Alias para compatibilidad con código existente
generar_excel_personalizado = generate_excel


# =============================================================================
# EJEMPLOS DE USO
# =============================================================================

if __name__ == "__main__":

    # ═══════════════════════════════════════════════════════════════════════════
    # TUTORIAL: Generar estructura vacía para horarios de grupos
    # ═══════════════════════════════════════════════════════════════════════════
    #
    # Este ejemplo muestra cómo generar un Excel con:
    #   - Hojas para cada grupo (horario + tabla de asignaturas vacía)
    #   - Hoja "Aulas" con fórmulas cruzadas
    #   - Fórmulas y formato condicional configurados
    #
    # Solo necesitas configurar los nombres de los grupos.

    # ── Configuración ─────────────────────────────────────────────────
    GRUPOS = ["D111", "D211", "C111" , "M111"]  # ← Solo cambia esto
    DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    AULAS = [f"Aula {i}" for i in range(1, 10)] + ["Lab"]
    TURNOS = 6
    ROW_STEP = 3  # filas por turno (nombre turno + aula + separador)

    def crear_hoja_grupo_vacia(nombre_grupo: str) -> dict:
        """Crea una hoja de grupo con tablas vacías pero con fórmulas."""

        # Calcular dimensiones dinámicas
        total_filas_horario = TURNOS * ROW_STEP + 1  # +1 por separador tras turno 3
        fin_horario_fila = 3 + total_filas_horario   # fila donde termina horario

        # Tabla de asignaturas vacía (solo headers, sin datos)
        filas_asignaturas = 1  # mínimo 1 fila vacía
        fin_asig_fila = 3 + filas_asignaturas
        rango_asignaturas = f"I4:M{fin_asig_fila}"
        rango_asig_abrev = f"$I4:I{fin_asig_fila}"

        # Tabla de aulas lateral
        filas_aulas = max(1, min(len(AULAS), total_filas_horario))
        fin_aulas_fila = 3 + filas_aulas
        rango_aulas = f"O4:O{fin_aulas_fila}"
        rango_aulas_cf = f"$O4:O{fin_aulas_fila}"

        # Construir datos de la hoja
        # Fila 1: título grupo, Fila 2: vacía, Fila 3: headers
        fila_headers = ["", "", *DIAS, "", "Abrev", "Asignaturas", "Frec", "Faltan", "Asignadas", "", "Aulas"]

        # Generar filas del horario (Turno 1, Turno 2, etc. con separadores)
        datos = [
            ["Grupo", nombre_grupo] + [""] * 13,
            [""] * 15,
            fila_headers,
        ]

        # Generar filas del horario con estructura: [Turno N, "", ""] x 3 veces + separador
        filas_por_turno = [
            [f"Turno {t}", "", ""] if s == 0 else ["", "", ""]
            for t in range(1, TURNOS + 1)
            for s in range(ROW_STEP)
        ]
        # Insertar separador después del turno 3 (fila 9, antes de Turno 4)
        filas_por_turno.insert(9, ["", "", ""])

        for i, (etiqueta, _, _) in enumerate(filas_por_turno):
            # Fila vacía del horario + asignaturas vacía + aula
            fila = [
                "",           # A
                etiqueta,     # B
                "", "", "", "", "",  # C-G: horario vacío (5 días)
                "",           # H
                "", "", "", "", "",  # I-M: asignaturas vacía
                "",           # N
                AULAS[i] if i < len(AULAS) else "",  # O: aula
            ]
            datos.append(fila)

        # Fórmulas
        formulas = [
            # Totales asignaturas
            {"row": fin_asig_fila + 1, "col": 12, "value": "Total:"},
            {"row": fin_asig_fila + 1, "col": 13, "value": f"=COUNTA(I4:I{fin_asig_fila})"},
            {"row": fin_asig_fila + 2, "col": 12, "value": "Σ Frec:"},
            {"row": fin_asig_fila + 2, "col": 13, "value": f"=SUM(K4:K{fin_asig_fila})"},
            # Totales aulas
            {"row": fin_aulas_fila + 1, "col": 14, "value": "Total:"},
            {"row": fin_aulas_fila + 1, "col": 15, "value": f"=COUNTA(O4:O{fin_aulas_fila})"},
            # Ocupados horario
            {"row": fin_horario_fila + 1, "col": 6, "value": "Ocupados:"},
            {"row": fin_horario_fila + 1, "col": 7, "value": f"=COUNTA(C4:G{fin_horario_fila})/{ROW_STEP}"},
        ]

        # Calcular posición exacta del separador en el Excel
        # Turno 1: filas 4-6, Turno 2: 7-9, Turno 3: 10-12, Separador: 13, Turno 4: 14-16...
        fila_separador = 4 + (3 * ROW_STEP)  # fila 13

        # Rangos del horario divididos por el separador (para formato condicional)
        rango_horario_pre = f"C4:G{fila_separador - 1}"   # C4:G12 (Turnos 1-3)
        rango_horario_post = f"C{fila_separador + 1}:G{fin_horario_fila}"  # C14:G22 (Turnos 4-6)
        rangos_horario_cf = [rango_horario_pre, rango_horario_post]

        # Rangos para merges de turnos (columna B)
        merges = []
        fila = 4
        for turno in range(1, TURNOS + 1):
            inicio = fila
            fin = fila + ROW_STEP - 1
            merges.append(f"B{inicio}:B{fin}")
            fila += ROW_STEP
            if turno == 3:
                fila += 1  # saltar fila separadora después de Turno 3

        # Formato condicional - aplicar a cada rango del horario (pre y post separador)
        reglas_cf = []
        for rango_h in rangos_horario_cf:
            reglas_cf.extend([
                # Filas pares del horario: validar que abrev exista en asignaturas
                {
                    "tipo": "filas_pares",
                    "rango": rango_h,
                    "formula": f'AND({{celda}}<>"", COUNTIF({rango_asig_abrev},{{celda}})=0)',
                    "color": "F4A460",
                    "row_step": ROW_STEP,
                    "row_start_offset": 0,
                },
                # Filas impares del horario: validar que aula exista en catálogo
                {
                    "tipo": "filas_impares",
                    "rango": rango_h,
                    "formula": f'AND({{celda}}<>"", COUNTIF({rango_aulas_cf},{{celda}})=0)',
                    "color": "FFD700",
                    "row_step": ROW_STEP,
                    "row_start_offset": 1,
                },
                # Alerta roja si fila par tiene valor pero siguiente (impar) está vacía
                {
                    "tipo": "pares_con_siguiente",
                    "rango": rango_h,
                    "formula": 'AND({celda}<>"", {celda_siguiente}="")',
                    "color": "FF0000",
                    "row_step": ROW_STEP,
                    "next_offset": 1,
                    "aplicar_a": "siguiente",
                },
            ])

        return {
            "title": nombre_grupo,
            "data": datos,
            "column_widths": {i: 14 for i in range(1, 16)},
            "range_styles": [
                {"range": f"I3:I{fin_asig_fila}", "style": {"bg_color": "A9D18E"}},  # verde asignaturas
                {"range": f"B4:B{fin_horario_fila}", "style": {"bg_color": "F4CCCC"}},  # rojo turnos
            ],
            "table_ranges": [
                "B3:G3",
                # Rangos de los turnos en columna B (las celdas fusionadas)
                f"B4:B{fila_separador - 1}",      # B4:B12 (Turnos 1-3)
                f"B{fila_separador + 1}:B{fin_horario_fila}",  # B14:B22 (Turnos 4-6)
                rango_horario_pre,   # C4:G12 (Turnos 1-3)
                rango_horario_post,  # C14:G22 (Turnos 4-6)
                f"B{fila_separador}:G{fila_separador}",  # B13:G13 Separador
                "I3:M3",
                rango_asignaturas,
                rango_aulas,
            ],
            "table_block_sizes": [
                {"range": "B3:G3", "row_step": 1, "col_step": 1},
                {"range": f"B4:B{fila_separador - 1}", "row_step": ROW_STEP, "col_step": 1},  # Turnos col B (1-3)
                {"range": f"B{fila_separador + 1}:B{fin_horario_fila}", "row_step": ROW_STEP, "col_step": 1},  # Turnos col B (4-6)
                {"range": rango_horario_pre, "row_step": ROW_STEP, "col_step": 1},
                {"range": rango_horario_post, "row_step": ROW_STEP, "col_step": 1},
                {"range": f"B{fila_separador}:G{fila_separador}", "row_step": 1, "col_step": 1},
                {"range": "I3:M3", "row_step": 1, "col_step": 1},
                {"range": rango_asignaturas, "row_step": 1, "col_step": 1},
                {"range": rango_aulas, "row_step": 1, "col_step": 1},
            ],
            "merge_ranges": merges,
            "table_borders": True,
            "border_color": "4F81BD",
            "border_style": "medium",
            "formulas": formulas,
            "conditional_format_rules": reglas_cf,
        }


    def crear_hoja_aulas(grupos: list[str]) -> dict:
        """Crea la hoja Aulas con fórmulas cruzadas a los grupos."""
        # Estructura: 5 bloques de días, cada uno con header + 6 turnos
        datos = []
        turnos_labels = ["1ro", "2do", "3ro", "4to", "5to", "6to"]

        for dia_idx, dia in enumerate(DIAS):
            # Fila vacía separadora (excepto antes del primer bloque)
            if dia_idx > 0:
                datos.append([""] * 12)
            # Header del día con nombres de aulas
            datos.append(["", dia] + AULAS)
            # 6 filas de turnos vacías
            for t in range(TURNOS):
                datos.append(["", turnos_labels[t]] + [""] * 10)

        return {
            "title": "Aulas",
            "data": datos,
            "column_widths": {i: 12 for i in range(1, 13)},
            "range_styles": [
                {"range": "B1:L1", "style": {"bg_color": "E6B8AF"}},
            ],
            "table_borders": True,
            "border_color": "B3B3B3",
            "border_style": "thick",
            # Las fórmulas cruzadas se inyectan vía fernando_formulas
            # (requiere lógica adicional para mapear celdas)
        }


    # ── Generar configuración completa ─────────────────────────────────
    sheets = [crear_hoja_grupo_vacia(g) for g in GRUPOS]
    sheets.append(crear_hoja_aulas(GRUPOS))

    config = {"sheets": sheets}

    # ── Ejecutar ───────────────────────────────────────────────────────
    generate_excel(config, "tutorial_estructura_vacia.xlsx")
    print("✅ Archivo 'tutorial_estructura_vacia.xlsx' generado.")
    print(f"   Grupos: {', '.join(GRUPOS)}")
    print("   Cada grupo tiene: horario vacío, tabla asignaturas vacía, lista de aulas")
    print("   Hoja 'Aulas': tabla cruzada (lista para fórmulas)")
