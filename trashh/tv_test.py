#!/usr/bin/env python3
import openpyxl

# DATA
tv_lunes_data = [["16:00", "16:05", 5, "EL TIEMPO Y LA MEMORIA", "informativo", "adulto"], ["16:05", "16:10", 5, "COORDENADAS", "informativo", "adulto"], ["16:10", "17:00", 50, "REVISTA HOLA HABANA", "revista", "adulto"], ["17:00", "17:30", 30, "DÉCADAS MILAGROSAS", "musical", "adulto"], ["17:30", "18:00", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["18:00", "18:30", 30, "POWER RANGERS", "infantil", "infantil"], ["18:30", "18:55", 25, "CINECITO EN TV", "cine", "infantil"], ["18:55", "19:00", 5, "COORDENADAS INFANTILES", "informativo", "infantil"], ["19:00", "19:30", 30, "VE Y MIRA", "cine", "adulto"], ["19:30", "20:00", 30, "MÚSICA HABANA", "musical", "juvenil"], ["20:00", "20:30", 30, "HABANA COLECCIÓN", "cultural", "adulto"], ["20:30", "21:30", 60, "MÚSICA SÍ", "musical", "adulto"], ["21:30", "21:45", 15, "D DISEÑO", "cultural", "adulto"], ["21:45", "22:00", 15, "SIN PUNTOS SUSPENSIVOS", "entrevista", "adulto"], ["22:00", "22:45", 45, "NOVELA “LA NIETA ELEGIDA”", "ficción", "adulto"], ["22:45", "23:15", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["23:15", "00:00", 45, "SERIE “CRÍMENES MAYORES”", "ficción", "adulto"]]
tv_martes_data = [["19:30", "20:00", 30, "TODO POP", "musical", "juvenil"], ["20:00", "20:30", 30, "BREVES ESTACIONES", "cultural", "adulto"], ["20:30", "20:45", 15, "SALUDARTE", "salud", "adulto"], ["20:45", "21:00", 15, "SECUENCIA", "cultural", "adulto"], ["21:00", "21:30", 30, "RITMO CLIP", "musical", "juvenil"], ["21:30", "22:00", 30, "TRIANGULO DE LA CONFIANZA", "entrevista", "adulto"], ["22:00", "22:45", 45, "NOVELA “LA NIETA ELEGIDA”", "ficción", "adulto"], ["22:45", "23:15", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["23:15", "00:00", 45, "SERIE “CRÍMENES MAYORES”", "ficción", "adulto"]]
tv_miércoles_data = [["19:45", "20:00", 15, "COSAS DEL CINE", "cine", "adulto"], ["20:00", "20:30", 30, "RITMO CLIP", "musical", "juvenil"], ["20:30", "21:00", 30, "BANDA SONORA JUVENIL", "cultural", "juvenil"], ["21:00", "21:30", 30, "VE Y MIRA", "cine", "adulto"], ["21:30", "22:00", 30, "MÚSICA DEL MUNDO", "musical", "adulto"], ["22:00", "22:30", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["22:30", "23:47", 77, "CINEMA HABANA", "cine", "adulto"]]
tv_jueves_data = [["16:00", "16:05", 5, "EL TIEMPO Y LA MEMORIA", "informativo", "adulto"], ["16:05", "16:10", 5, "COORDENADAS", "informativo", "adulto"], ["16:10", "17:00", 50, "REVISTA HOLA HABANA", "revista", "adulto"], ["17:00", "17:15", 15, "SALUDARTE", "salud", "adulto"], ["17:15", "17:30", 15, "SECUENCIA", "cultural", "adulto"], ["17:30", "18:00", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["18:00", "18:30", 30, "CAZADOR DE TROLES", "infantil", "infantil"], ["18:30", "18:45", 15, "EL MUNDO DE CRAIG", "animacion", "infantil"], ["18:45", "19:30", 45, "ÁNIMA", "animacion", "juvenil"], ["19:30", "20:00", 30, "PAPEL EN BLANCO", "cultural", "adulto"], ["20:00", "20:30", 30, "MÚSICA DEL MUNDO", "musical", "adulto"], ["20:30", "20:45", 15, "TRAVESÍA", "cultural", "adulto"], ["20:45", "21:00", 15, "DONDE VA LA HABANA", "cultural", "adulto"], ["21:00", "21:30", 30, "BANDA SONORA", "musical", "adulto"], ["21:30", "22:00", 30, "ESTA ES MI PEÑA", "musical", "adulto"], ["22:00", "22:45", 45, "NOVELA “LA NIETA ELEGIDA”", "ficción", "adulto"], ["22:45", "23:15", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["23:15", "00:00", 45, "SERIE “CRÍMENES MAYORES”", "ficción", "adulto"]]
tv_viernes_data = [["16:30", "17:30", 60, "GEN HABANERO", "documental", "adulto"], ["17:30", "18:00", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["18:00", "19:55", 115, "TIENE QUE VER", "cine", "infantil"], ["19:55", "20:00", 5, "COORDENADAS INFANTILES", "informativo", "infantil"], ["20:00", "20:30", 30, "ALGO ENTRE MANOS", "cultural", "adulto"], ["20:30", "21:00", 30, "BREVES ESTACIONES", "cultural", "adulto"], ["21:00", "21:30", 30, "DÉCADAS MILAGROSAS", "musical", "adulto"], ["21:30", "21:45", 15, "SIN PUNTOS SUSPENSIVOS", "entrevista", "adulto"], ["21:45", "22:00", 15, "YO BAILO", "musical", "toda-la-familia"], ["22:00", "22:45", 45, "NOVELA “LA NIETA ELEGIDA”", "ficción", "adulto"], ["22:45", "23:15", 30, "HABANA NOTICIARIO", "informativo", "adulto"], ["23:15", "00:00", 45, "SERIE “CRÍMENES MAYORES”", "ficción", "adulto"]]
tv_sábado_data = [["16:00", "16:05", 5, "EL TIEMPO Y LA MEMORIA", "informativo", "adulto"], ["16:05", "16:10", 5, "COORDENADAS", "informativo", "adulto"], ["16:10", "16:45", 35, "CUENTAS VERDE LIMÓN", "infantil", "infantil"], ["16:45", "17:15", 30, "BANDA SONORA JUVENIL", "cine", "juvenil"], ["17:15", "17:30", 15, "TRAVESÍA", "cultural", "adulto"], ["17:30", "18:00", 30, "TODO POP", "musical", "juvenil"], ["18:00", "18:45", 45, "SERIE JUVENIL “SABRINA, LA BRUJA ADOLESCENTE”", "ficción", "juvenil"], ["18:45", "19:30", 45, "LIBRE ACCESO", "informativo", "adulto"], ["19:30", "20:00", 30, "JUGADA PERFECTA", "deporte", "toda-la-familia"], ["20:00", "20:15", 15, "QUE LA MÚSICA NO FALTE", "musical", "adulto"], ["20:15", "20:30", 15, "COSAS DEL CINE", "cine", "adulto"], ["20:30", "21:00", 30, "MÚSICA HABANA", "musical", "adulto"], ["21:00", "22:30", 90, "X DISTANTE", "animacion", "juvenil"], ["22:30", "00:00", 90, "MÚSICA SI", "musical", "adulto"]]
tv_domingo_data = [["02:00", "02:05", 5, "EL TIEMPO Y LA MEMORIA", "informativo", "adulto"], ["02:05", "02:10", 5, "COORDENADAS", "informativo", "adulto"], ["02:10", "17:00", 890, "CANAL HABANA DEPORTES", "deporte", "toda-la-familia"], ["17:00", "18:00", 60, "LATINOS", "musical", "adulto"], ["18:00", "18:30", 30, "PAPEL EN BLANCO", "cultural", "adulto"], ["18:30", "19:00", 30, "ALGO ENTRE MANOS", "cultural", "adulto"], ["19:00", "19:15", 15, "VERDE HABANA", "documental", "toda-la-familia"], ["19:15", "19:30", 15, "GEN HABANERO", "documental", "adulto"], ["19:30", "20:00", 30, "TRIANGULO DE LA CONFIANZA", "entrevista", "adulto"], ["20:00", "20:30", 30, "BANDA SONORA", "musical", "adulto"], ["20:30", "23:55", 205, "CINE +", "cine", "adulto"]]

def generar():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Canal Habana"

    # Lunes
    for r,d in enumerate(tv_lunes_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)
    # Martes
    for r,d in enumerate(tv_martes_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)
    # Miércoles
    for r,d in enumerate(tv_miércoles_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)
    # Jueves
    for r,d in enumerate(tv_jueves_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)
    # Viernes
    for r,d in enumerate(tv_viernes_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)
    # Sábado
    for r,d in enumerate(tv_sábado_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)
    # Domingo
    for r,d in enumerate(tv_domingo_data,1):
        for c,v in enumerate(d,1):
            ws.cell(r,c,v)

    return wb

def main():
    wb = generar()
    wb.save('Canal Habana.xlsx');print('OK: Canal Habana.xlsx')

if __name__=='__main__':main()
