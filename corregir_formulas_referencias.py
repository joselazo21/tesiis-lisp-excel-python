"""
Script para corregir referencias de hojas en fórmulas Excel
Corrige el uso incorrecto de punto (.) por exclamación (!) en referencias a hojas
"""

import openpyxl
import re
import sys
from pathlib import Path


def corregir_formula(formula: str) -> str:
    """
    Corrige referencias de hojas en una fórmula Excel.
    
    Convierte referencias como $C111.$C$7 a C111!$C$7
    Convierte referencias como c112!$c$7 a C112!$C$7 (normaliza mayúsculas)
    
    Args:
        formula: Fórmula Excel con posibles errores
        
    Returns:
        Fórmula corregida
    """
    if not formula or not formula.startswith('='):
        return formula
    
    # Patrón para detectar referencias con punto: $C111.$C$7 o C111.$C$7
    # Captura: (nombre_hoja).(referencia_celda)
    patron_punto = r'\$?([A-Za-z]\d+)\.\$?([A-Z]+\$?\d+)'
    
    def reemplazo(match):
        hoja = match.group(1).upper()  # Normalizar a mayúsculas
        celda = match.group(2)
        return f'{hoja}!${celda}' if not celda.startswith('$') else f'{hoja}!{celda}'
    
    formula_corregida = re.sub(patron_punto, reemplazo, formula)
    
    # Normalizar nombres de hojas en referencias correctas (c112!$c$7 -> C112!$C$7)
    patron_minusculas = r'([a-z]\d+)!'
    
    def normalizar_hoja(match):
        return match.group(1).upper() + '!'
    
    formula_corregida = re.sub(patron_minusculas, normalizar_hoja, formula_corregida)
    
    return formula_corregida


def corregir_archivo(ruta_archivo: str, crear_backup: bool = True) -> dict:
    """
    Corrige todas las fórmulas en un archivo Excel.
    
    Args:
        ruta_archivo: Ruta al archivo Excel
        crear_backup: Si True, crea un backup del archivo original
        
    Returns:
        Dict con estadísticas de corrección
    """
    ruta = Path(ruta_archivo)
    
    if not ruta.exists():
        return {'error': f'Archivo no encontrado: {ruta_archivo}'}
    
    # Crear backup si se solicita
    if crear_backup:
        backup_path = ruta.with_suffix('.backup.xlsx')
        import shutil
        shutil.copy2(ruta, backup_path)
        print(f"✓ Backup creado: {backup_path.name}")
    
    try:
        wb = openpyxl.load_workbook(ruta)
        stats = {
            'archivo': ruta.name,
            'hojas_procesadas': 0,
            'celdas_corregidas': 0,
            'formulas_corregidas': []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            hoja_modificada = False
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_original = cell.value
                        formula_corregida = corregir_formula(formula_original)
                        
                        if formula_original != formula_corregida:
                            cell.value = formula_corregida
                            stats['celdas_corregidas'] += 1
                            hoja_modificada = True
                            
                            stats['formulas_corregidas'].append({
                                'hoja': sheet_name,
                                'celda': cell.coordinate,
                                'original': formula_original[:80],
                                'corregida': formula_corregida[:80]
                            })
            
            if hoja_modificada:
                stats['hojas_procesadas'] += 1
        
        # Guardar solo si hubo cambios
        if stats['celdas_corregidas'] > 0:
            wb.save(ruta)
            print(f"✓ Archivo guardado: {ruta.name}")
        else:
            print(f"ℹ No se encontraron fórmulas para corregir en: {ruta.name}")
        
        wb.close()
        return stats
        
    except Exception as e:
        return {'error': f'Error procesando {ruta.name}: {str(e)}'}


def main():
    """Función principal para corregir archivos."""
    
    archivos_a_corregir = [
        'Aulas_Con_Formulas_Fernando.xlsx',
        'prueba_formulas_fernando_es.xlsx',
        'prueba_formulas_fernando_en.xlsx',
        'prueba_formulas_fernando.xlsx'
    ]
    
    print("=" * 70)
    print("CORRECTOR DE REFERENCIAS EN FÓRMULAS EXCEL")
    print("=" * 70)
    print()
    
    resultados = []
    
    for archivo in archivos_a_corregir:
        print(f"\nProcesando: {archivo}")
        print("-" * 70)
        
        stats = corregir_archivo(archivo, crear_backup=True)
        resultados.append(stats)
        
        if 'error' in stats:
            print(f"❌ {stats['error']}")
        else:
            print(f"  Hojas procesadas: {stats['hojas_procesadas']}")
            print(f"  Celdas corregidas: {stats['celdas_corregidas']}")
            
            if stats['celdas_corregidas'] > 0:
                print(f"\n  Ejemplos de correcciones:")
                for correccion in stats['formulas_corregidas'][:3]:
                    print(f"    • {correccion['hoja']}!{correccion['celda']}")
                    print(f"      ANTES: {correccion['original']}")
                    print(f"      DESPUÉS: {correccion['corregida']}")
    
    # Resumen final
    print("\n" + "=" * 70)
    print("RESUMEN")
    print("=" * 70)
    
    total_corregidas = sum(r.get('celdas_corregidas', 0) for r in resultados)
    total_errores = sum(1 for r in resultados if 'error' in r)
    
    print(f"Total de celdas corregidas: {total_corregidas}")
    print(f"Archivos con errores: {total_errores}")
    print(f"\n✓ Proceso completado")


if __name__ == '__main__':
    main()
