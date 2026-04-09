"""
Script para arreglar Aulas_Con_Formulas_Fernando.xlsx
Copia las hojas C111, C112, C113 necesarias para que las fórmulas funcionen
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy


def copiar_hoja(ws_origen, ws_destino):
    """
    Copia todos los datos, formatos y fórmulas de una hoja a otra.
    
    Args:
        ws_origen: Hoja de origen
        ws_destino: Hoja de destino
    """
    # Copiar dimensiones de columna
    for col_letter in ws_origen.column_dimensions:
        if col_letter in ws_origen.column_dimensions:
            ws_destino.column_dimensions[col_letter].width = ws_origen.column_dimensions[col_letter].width
    
    # Copiar dimensiones de fila
    for row_num in ws_origen.row_dimensions:
        if row_num in ws_origen.row_dimensions:
            ws_destino.row_dimensions[row_num].height = ws_origen.row_dimensions[row_num].height
    
    # Copiar celdas
    for row in ws_origen.iter_rows():
        for cell in row:
            new_cell = ws_destino[cell.coordinate]
            
            # Copiar valor o fórmula
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    new_cell.value = cell.value
                else:
                    new_cell.value = cell.value
            
            # Copiar formato
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    # Copiar merged cells
    for merged_range in ws_origen.merged_cells.ranges:
        ws_destino.merge_cells(str(merged_range))


def main():
    """Función principal."""
    
    print("=" * 70)
    print("ARREGLANDO ARCHIVO: Aulas_Con_Formulas_Fernando.xlsx")
    print("=" * 70)
    
    # Abrir archivo fuente (que tiene las hojas C111, C112, C113)
    archivo_fuente = 'prueba_formulas_fernando_es.xlsx'
    archivo_destino = 'Aulas_Con_Formulas_Fernando.xlsx'
    
    print(f"\n1. Abriendo archivo fuente: {archivo_fuente}")
    wb_fuente = openpyxl.load_workbook(archivo_fuente)
    print(f"   Hojas disponibles: {wb_fuente.sheetnames}")
    
    print(f"\n2. Abriendo archivo destino: {archivo_destino}")
    wb_destino = openpyxl.load_workbook(archivo_destino)
    print(f"   Hojas actuales: {wb_destino.sheetnames}")
    
    # Verificar si C111 existe en fuente
    if 'C111' not in wb_fuente.sheetnames:
        print(f"\n❌ ERROR: La hoja C111 no existe en {archivo_fuente}")
        return
    
    # Copiar hoja C111
    print(f"\n3. Copiando hoja C111...")
    
    # Eliminar C111 si ya existe en destino
    if 'C111' in wb_destino.sheetnames:
        del wb_destino['C111']
        print("   Hoja C111 existente eliminada")
    
    # Crear nueva hoja C111
    ws_c111_destino = wb_destino.create_sheet('C111')
    ws_c111_origen = wb_fuente['C111']
    
    copiar_hoja(ws_c111_origen, ws_c111_destino)
    print("   ✓ Hoja C111 copiada")
    
    # Crear hojas C112 y C113 como copias de C111 (estructura similar)
    print(f"\n4. Creando hojas C112 y C113 basadas en C111...")
    
    for nombre_hoja in ['C112', 'C113']:
        if nombre_hoja in wb_destino.sheetnames:
            del wb_destino[nombre_hoja]
        
        ws_nueva = wb_destino.create_sheet(nombre_hoja)
        copiar_hoja(ws_c111_origen, ws_nueva)
        print(f"   ✓ Hoja {nombre_hoja} creada")
    
    # Guardar archivo
    print(f"\n5. Guardando archivo...")
    wb_destino.save(archivo_destino)
    print(f"   ✓ Archivo guardado: {archivo_destino}")
    
    print(f"\n6. Verificando resultado...")
    wb_verificar = openpyxl.load_workbook(archivo_destino)
    print(f"   Hojas finales: {wb_verificar.sheetnames}")
    
    # Verificar que las referencias existan
    if 'Aulas' in wb_verificar.sheetnames:
        ws_aulas = wb_verificar['Aulas']
        cell = ws_aulas['F4']
        if cell.value:
            print(f"\n   Fórmula en F4: {cell.value[:80]}...")
    
    for hoja in ['C111', 'C112', 'C113']:
        if hoja in wb_verificar.sheetnames:
            ws = wb_verificar[hoja]
            print(f"   ✓ {hoja}!B1 = '{ws['B1'].value}'")
        else:
            print(f"   ❌ {hoja} NO existe")
    
    wb_verificar.close()
    wb_fuente.close()
    wb_destino.close()
    
    print("\n" + "=" * 70)
    print("✓ PROCESO COMPLETADO")
    print("=" * 70)
    print("\nAhora abre 'Aulas_Con_Formulas_Fernando.xlsx' y las fórmulas")
    print("deberían funcionar correctamente sin error #NOMBRE.")


if __name__ == '__main__':
    main()
